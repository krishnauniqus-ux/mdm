"""Data Profiling Component - AI-Powered Dynamic Validation with Azure OpenAI"""

import streamlit as st
import pandas as pd
import io
import re
import json
import unicodedata
import os
from datetime import datetime
from collections import Counter, defaultdict
import time
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from typing import Dict, List, Any, Optional
import hashlib
import logging

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('data_profiler_validation.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Azure OpenAI imports
try:
    from openai import AzureOpenAI
except ImportError:
    st.error("Please install: pip install openai")

from state.session import st, show_toast
from core.profiler import DataProfilerEngine

# ==========================================
# AZURE OPENAI CONFIGURATION
# ==========================================

class AzureOpenAIConfig:
    """Azure OpenAI Configuration - Set these in environment variables or secrets"""
    AZURE_OPENAI_ENDPOINT = st.secrets.get("AZURE_OPENAI_ENDPOINT")
    AZURE_OPENAI_KEY = st.secrets.get("AZURE_OPENAI_KEY")
    AZURE_OPENAI_DEPLOYMENT = st.secrets.get("AZURE_OPENAI_DEPLOYMENT")
    AZURE_OPENAI_API_VERSION = st.secrets.get("AZURE_OPENAI_API_VERSION")
    
    # Rate limiting configuration (adjust based on your Azure tier)
    # Standard: 60 requests/minute, 40K tokens/minute
    # Premium: Higher limits
    MAX_REQUESTS_PER_MINUTE = int(st.secrets.get("AZURE_OPENAI_MAX_RPM", 60))
    MAX_TOKENS_PER_MINUTE = int(st.secrets.get("AZURE_OPENAI_MAX_TPM", 40000))
     
    @classmethod
    def validate(cls):
        missing = []
        if not cls.AZURE_OPENAI_ENDPOINT:
            missing.append("AZURE_OPENAI_ENDPOINT")
        if not cls.AZURE_OPENAI_KEY:
            missing.append("AZURE_OPENAI_KEY")
        if not cls.AZURE_OPENAI_DEPLOYMENT:
            missing.append("AZURE_OPENAI_DEPLOYMENT")
        return missing

# ==========================================
# METADATA EXTRACTION UTILITIES
# ==========================================

def extract_column_metadata(column_name: str) -> Dict[str, Any]:
    """Extract metadata from column names like VARCHAR2(360), CHAR(19), etc.
    
    Returns dict with:
    - max_length: int or None
    - data_type_hint: str or None (VARCHAR, CHAR, NUMBER, etc.)
    - original_name: str
    """
    metadata = {
        'max_length': None,
        'data_type_hint': None,
        'original_name': column_name
    }
    
    # Pattern 1: VARCHAR2(360), VARCHAR(26), CHAR(19), CHARACTERS(21), etc.
    # Matches: VARCHAR2 (360 CHAR), VARCHAR(26), CHAR(19), CHARACTERS (21), etc.
    pattern1 = r'(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?'
    match = re.search(pattern1, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['max_length'] = int(match.group(2))
        logger.info(f"Extracted metadata from '{column_name}': Type={metadata['data_type_hint']}, MaxLength={metadata['max_length']}")
        return metadata
    
    # Pattern 2: NUMBER(10,2), DECIMAL(15,2), etc.
    pattern2 = r'(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?'
    match = re.search(pattern2, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['max_length'] = int(match.group(2))  # precision
        logger.info(f"Extracted metadata from '{column_name}': Type={metadata['data_type_hint']}, Precision={metadata['max_length']}")
        return metadata
    
    # Pattern 3: Just numbers in parentheses like "Name (360)" or "Code (19)"
    pattern3 = r'\((\d+)\)'
    match = re.search(pattern3, column_name)
    if match:
        metadata['max_length'] = int(match.group(1))
        logger.info(f"Extracted metadata from '{column_name}': MaxLength={metadata['max_length']}")
        return metadata
    
    return metadata


def extract_excel_cell_notes(uploaded_file, sheet_name: str = None) -> Dict[str, str]:
    """Extract cell notes/comments from Excel file
    
    Returns dict mapping column names to their notes/comments
    """
    try:
        import openpyxl
        
        # Reset file pointer
        uploaded_file.seek(0)
        
        # Load workbook with data_only=False to get comments
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        
        # Get the specified sheet or first sheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        column_notes = {}
        
        # Check first row (header row) for comments
        for cell in ws[1]:
            if cell.comment:
                column_name = str(cell.value)
                comment_text = cell.comment.text
                column_notes[column_name] = comment_text
                logger.info(f"Found Excel comment for column '{column_name}': {comment_text[:100]}")
        
        # Also check second row for metadata/rules
        if ws.max_row >= 2:
            for idx, cell in enumerate(ws[2], start=1):
                header_cell = ws.cell(row=1, column=idx)
                column_name = str(header_cell.value)
                
                # Check if second row contains metadata (like VARCHAR2(360))
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    # Check if it looks like a data type definition
                    if re.search(r'(VARCHAR|CHAR|NUMBER|DECIMAL|DATE|TIMESTAMP)', cell_value, re.IGNORECASE):
                        if column_name not in column_notes:
                            column_notes[column_name] = cell_value
                        else:
                            column_notes[column_name] += f" | {cell_value}"
                        logger.info(f"Found metadata in row 2 for column '{column_name}': {cell_value}")
        
        wb.close()
        return column_notes
        
    except ImportError:
        logger.warning("openpyxl not installed. Cannot extract Excel comments. Install with: pip install openpyxl")
        return {}
    except Exception as e:
        logger.error(f"Error extracting Excel comments: {str(e)}")
        return {}


# ==========================================
# AI-POWERED VALIDATION ENGINE
# ==========================================

class AIValidationEngine:
    """Enterprise-grade AI-powered validation rule generator using Azure OpenAI"""
    
    # Strict DQ Dimensions as per enterprise standards + Client Custom Dimension
    DQ_DIMENSIONS = [
        "Accuracy", "Completeness", "Consistency", "Validity", 
        "Uniqueness", "Timeliness", "Integrity", "Conformity",
        "Reliability", "Relevance", "Precision", "Accessibility",
        "Character Length"  # NEW: For client-specified character limits
    ]
    
    def __init__(self):
        self.client = None
        self.cache = {}  # Cache for validation rules to avoid repeated API calls
        self._initialize_client()
    
    def _initialize_client(self):
        """Initialize Azure OpenAI client"""
        try:
            missing = AzureOpenAIConfig.validate()
            if missing:
                st.warning(f"Azure OpenAI not configured. Missing: {', '.join(missing)}")
                return
            
            self.client = AzureOpenAI(
                azure_endpoint=AzureOpenAIConfig.AZURE_OPENAI_ENDPOINT,
                api_key=AzureOpenAIConfig.AZURE_OPENAI_KEY,
                api_version=AzureOpenAIConfig.AZURE_OPENAI_API_VERSION
            )
        except Exception as e:
            st.error(f"Failed to initialize Azure OpenAI: {str(e)}")
    
    def _get_cache_key(self, column_name: str, sample_data: List[str], data_type: str) -> str:
        """Generate cache key for column analysis"""
        data_hash = hashlib.md5(str(sample_data[:50]).encode()).hexdigest()
        return f"{column_name}_{data_type}_{data_hash}"
    
    def _call_azure_openai(self, messages: List[Dict], temperature: float = 0.1) -> Optional[str]:
        """Make Azure OpenAI API call with error handling"""
        if not self.client:
            return None
        
        try:
            response = self.client.chat.completions.create(
                model=AzureOpenAIConfig.AZURE_OPENAI_DEPLOYMENT,
                messages=messages,
                temperature=temperature,
                max_completion_tokens=4000  # Increased from 2000 to handle more rules
            )
            return response.choices[0].message.content
        except Exception as e:
            # Check if it's a token limit error
            error_msg = str(e)
            if "token" in error_msg.lower() or "length" in error_msg.lower():
                st.warning(f"Token limit reached for this column. Using fallback rules.")
                return None
            st.error(f"Azure OpenAI API Error: {str(e)}")
            return None
    
    def analyze_column_semantic_type(self, column_name: str, sample_data: List[str], 
                                    data_type: str, null_pct: float, unique_pct: float,
                                    metadata: Dict[str, Any] = None, excel_notes: str = None) -> Dict[str, Any]:
        """Use AI to determine the semantic type and validation rules for a column
        
        Args:
            column_name: Name of the column
            sample_data: Sample values from the column
            data_type: Pandas data type
            null_pct: Percentage of null values
            unique_pct: Percentage of unique values
            metadata: Extracted metadata (max_length, data_type_hint, etc.)
            excel_notes: Notes/comments from Excel file
        """
        
        cache_key = self._get_cache_key(column_name, sample_data, data_type)
        if cache_key in self.cache:
            return self.cache[cache_key]
        
        # Prepare sample data for AI analysis
        samples = sample_data[:100] if sample_data else []
        sample_str = json.dumps(samples, indent=2)
        
        # Build metadata section
        metadata_section = ""
        if metadata:
            metadata_section = "\nEXTRACTED METADATA FROM COLUMN NAME:\n"
            if metadata.get('max_length'):
                metadata_section += f"- Maximum Length: {metadata['max_length']} characters\n"
            if metadata.get('data_type_hint'):
                metadata_section += f"- Data Type Hint: {metadata['data_type_hint']}\n"
        
        if excel_notes:
            metadata_section += f"\nEXCEL NOTES/COMMENTS:\n{excel_notes}\n"
        
        # HUMAN READABLE RULE PROMPT WITH METADATA
        prompt = f"""Analyze this data column and generate human-readable validation rules.

COLUMN INFORMATION
==================
Column Name: {column_name}
Data Type: {data_type}
Sample Values: {sample_str}
Null Percentage: {null_pct:.1f}%
Unique Percentage: {unique_pct:.1f}%
{metadata_section}

IMPORTANT INSTRUCTIONS
======================
1. If METADATA shows a maximum length (e.g., VARCHAR2(360), CHAR(19)), you MUST create a "Conformity" rule for it
   Example: "Supplier Name must not exceed 360 characters"
   
2. If EXCEL NOTES contain validation rules or constraints, incorporate them into your rules

3. Even if the column has NO DATA (100% null), still generate meaningful rules based on:
   - Column name semantics
   - Metadata constraints
   - Excel notes
   - Industry standards for that field type

REQUIRED OUTPUT FORMAT
======================
Return a JSON object with this exact structure:

{{
  "business_field_name": "Human-friendly field name (e.g., 'Posting Date', 'Email Address', 'Mobile Number')",
  "rules": [
    {{
      "dimension": "One of: Accuracy, Completeness, Consistency, Validity, Uniqueness, Timeliness, Integrity, Conformity, Reliability, Relevance, Precision, Accessibility, Character Length",
      "rule_statement": "Human readable rule in format: [Field Name] + Must/Should + Business Condition. Example: 'Posting Date must not be future dated'"
    }}
  ]
}}

RULE WRITING GUIDELINES
=======================
Write rules in this exact format: [Field Name] + Must/Should + Business Condition

EXAMPLES BY FIELD TYPE:

Date Fields (Date of Birth, Posting Date, Invoice Date):
- Date of Birth must be a valid calendar date
- Date of Birth cannot be in the future
- Date of Birth cannot be more than 120 years in the past
- Posting Date must not be future dated
- Posting Date must be within the active financial period
- Invoice Date should not be blank

Email Fields (Email Address):
- Email Address must follow standard email format
- Email Address must contain @ symbol
- Email Address must contain valid domain name
- Email Address should not contain spaces
- Email Address must be unique for each customer

Phone Fields (Mobile Number, Phone):
- Mobile Number must contain only digits
- Mobile Number must have 10 digits
- Mobile Number must start with valid prefix (6,7,8,9)
- Mobile Number should not contain special characters

ID Fields (PAN, GST, Asset ID):
- PAN must follow government issued format
- PAN must be 10 characters
- PAN must be alphanumeric
- Asset ID must follow organization coding standard

Amount Fields (Amount, Cost, Price):
- Amount must be numeric
- Amount should not be negative
- Amount must have maximum 2 decimal places
- Amount should not exceed defined business limit

Text Fields (Description, Name):
- Description should not be blank
- Description should not contain only special characters
- Description should not contain generic values like "Test" or "NA"

REQUIREMENTS
============
1. Generate 1-3 meaningful rules per column
2. Use exact format: [Field Name] + Must/Should + Business Condition
3. Rules must be immediately understandable by business users
4. No technical jargon, regex, or code references
5. Each rule should be actionable and specific
6. Assign most appropriate DQ dimension

Return ONLY valid JSON, no markdown."""

        messages = [
            {"role": "system", "content": "You are a Business Data Analyst who writes clear, human-readable data quality rules for enterprise systems. Return only JSON."},
            {"role": "user", "content": prompt}
        ]
        
        response = self._call_azure_openai(messages, temperature=0.1)
        
        if response:
            try:
                # Clean response - remove markdown code blocks if present
                cleaned = re.sub(r'```json\s*|\s*```', '', response).strip()
                result = json.loads(cleaned)
                self.cache[cache_key] = result
                return result
            except json.JSONDecodeError as e:
                st.warning(f"Failed to parse AI response for {column_name}: {str(e)}")
        
        # Fallback to basic analysis if AI fails
        return self._fallback_analysis(column_name, sample_data, data_type, null_pct, unique_pct)
    
    def _fallback_analysis(self, column_name: str, sample_data: List[str], 
                          data_type: str, null_pct: float, unique_pct: float) -> Dict[str, Any]:
        """Basic fallback analysis when AI is unavailable"""
        rules = []
        field_name = column_name.replace('_', ' ').title()
        
        # Completeness rule for nulls
        if null_pct > 0:
            rules.append({
                "dimension": "Completeness",
                "rule_statement": f"{field_name} should not be blank or null"
            })
        
        # Uniqueness rule for near-unique
        if unique_pct >= 95:
            rules.append({
                "dimension": "Uniqueness",
                "rule_statement": f"{field_name} must be unique"
            })
        
        # Validity rule based on data type
        if 'date' in data_type.lower():
            rules.append({
                "dimension": "Validity",
                "rule_statement": f"{field_name} must be a valid calendar date"
            })
            rules.append({
                "dimension": "Timeliness",
                "rule_statement": f"{field_name} must not be future dated"
            })
        elif 'int' in data_type.lower() or 'float' in data_type.lower():
            rules.append({
                "dimension": "Validity",
                "rule_statement": f"{field_name} must be numeric"
            })
        elif 'object' in data_type.lower():
            rules.append({
                "dimension": "Validity",
                "rule_statement": f"{field_name} must contain valid text"
            })
        
        return {
            "business_field_name": field_name,
            "rules": rules
        }
    
    def generate_dynamic_rules(self, df: pd.DataFrame, column_name: str, 
                              profile: Any, excel_notes: Dict[str, str] = None) -> List[Dict[str, Any]]:
        """Generate comprehensive validation rules for a column using AI
        
        Args:
            df: DataFrame containing the data
            column_name: Name of the column to analyze
            profile: Column profile with statistics
            excel_notes: Dictionary of Excel notes/comments by column name
        """
        logger.info(f"Generating dynamic rules for column: {column_name}")
        
        # Extract metadata from column name
        metadata = extract_column_metadata(column_name)
        
        # Get Excel notes for this column
        notes = None
        if excel_notes and column_name in excel_notes:
            notes = excel_notes[column_name]
            logger.info(f"  Found Excel notes for {column_name}: {notes[:100]}")
        
        sample_data = df[column_name].dropna().astype(str).head(100).tolist()
        data_type = str(profile.dtype)
        null_pct = profile.null_percentage
        unique_pct = profile.unique_percentage
        
        logger.debug(f"  Sample data size: {len(sample_data)}")
        logger.debug(f"  Data type: {data_type}")
        logger.debug(f"  Null %: {null_pct:.1f}%, Unique %: {unique_pct:.1f}%")
        if metadata.get('max_length'):
            logger.info(f"  Extracted max length: {metadata['max_length']}")
        if metadata.get('data_type_hint'):
            logger.info(f"  Extracted data type hint: {metadata['data_type_hint']}")
        
        # Get AI analysis with metadata
        try:
            ai_analysis = self.analyze_column_semantic_type(
                column_name, sample_data, data_type, null_pct, unique_pct,
                metadata=metadata, excel_notes=notes
            )
            logger.info(f"  AI analysis completed for {column_name}")
        except Exception as e:
            logger.error(f"  AI analysis failed for {column_name}: {str(e)}")
            return []
        
        rules = []
        field_name = ai_analysis.get('business_field_name', column_name)
        
        logger.info(f"  Business field name: {field_name}")
        logger.info(f"  Rules from AI: {len(ai_analysis.get('rules', []))}")
        
        # Process AI-generated validation rules
        for idx, rule in enumerate(ai_analysis.get('rules', [])):
            # Validate dimension is in allowed list
            dimension = rule.get('dimension', 'Validity')
            if dimension not in self.DQ_DIMENSIONS:
                logger.warning(f"  Invalid dimension '{dimension}' for rule {idx+1}. Using 'Validity'")
                dimension = 'Validity'
            
            validation_rule = {
                'S.No': len(rules) + 1,
                'Column': column_name,
                'Business Field': field_name,
                'Dimension': dimension,
                'Data Quality Rule': rule.get('rule_statement', 'No rule specified'),
                'Source': 'AI Generated'  # Mark source
            }
            rules.append(validation_rule)
            logger.debug(f"  Rule {idx+1}: {dimension} - {rule.get('rule_statement', 'No rule')[:50]}...")
        
        logger.info(f"  Total rules generated for {column_name}: {len(rules)}")
        return rules
    
    def validate_data_against_rules(self, df: pd.DataFrame, rules: List[Dict]) -> pd.DataFrame:
        """Validate dataframe against rules and return results with examples"""
        validation_results = []
        
        for rule in rules:
            column = rule.get('Column')
            if column not in df.columns:
                continue
            
            rule_statement = rule.get('Data Quality Rule', '').lower()
            invalid_count = 0
            invalid_examples = []
            
            # Validation logic based on rule statement keywords
            try:
                if 'not be blank' in rule_statement or 'not be null' in rule_statement:
                    mask = df[column].isna()
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).tolist()
                
                elif 'unique' in rule_statement:
                    mask = df[column].duplicated(keep=False)
                    invalid_count = mask.sum()
                    dup_values = df[mask][column].value_counts().head(3)
                    invalid_examples = [f"{val} ({count} times)" for val, count in dup_values.items()]
                
                elif 'valid calendar date' in rule_statement or 'valid date' in rule_statement:
                    converted = pd.to_datetime(df[column], errors='coerce')
                    mask = converted.isna() & df[column].notna()
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                elif 'not be future dated' in rule_statement or 'not be in the future' in rule_statement:
                    try:
                        dates = pd.to_datetime(df[column], errors='coerce')
                        mask = dates > datetime.now()
                        invalid_count = mask.sum()
                        invalid_examples = df[mask][column].head(5).astype(str).tolist()
                    except:
                        invalid_count = 0
                
                elif 'numeric' in rule_statement:
                    converted = pd.to_numeric(df[column], errors='coerce')
                    mask = converted.isna() & df[column].notna()
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                elif 'not be negative' in rule_statement:
                    try:
                        numeric_vals = pd.to_numeric(df[column], errors='coerce')
                        mask = numeric_vals < 0
                        invalid_count = mask.sum()
                        invalid_examples = df[mask][column].head(5).astype(str).tolist()
                    except:
                        invalid_count = 0
                
                elif 'only digits' in rule_statement:
                    mask = ~df[column].astype(str).str.match(r'^\d+$')
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                elif 'email' in rule_statement:
                    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                    mask = ~df[column].astype(str).str.match(email_pattern)
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                elif 'alphanumeric' in rule_statement:
                    mask = ~df[column].astype(str).str.match(r'^[a-zA-Z0-9]+$')
                    invalid_count = mask.sum()
                    invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                # Character length validation
                elif 'maximum' in rule_statement and 'character' in rule_statement:
                    match = re.search(r'maximum (\d+) character', rule_statement)
                    if match:
                        max_len = int(match.group(1))
                        mask = df[column].astype(str).str.len() > max_len
                        invalid_count = mask.sum()
                        invalid_examples = df[mask][column].head(5).astype(str).tolist()
                
                else:
                    invalid_count = 0
                    invalid_examples = []
                    
            except Exception as e:
                invalid_count = 0
                invalid_examples = []
            
            # Format examples for display
            if invalid_count == 0:
                examples_str = "✓ All values valid - No issues found"
            elif invalid_examples:
                cleaned_examples = []
                for ex in invalid_examples:
                    ex_str = str(ex).strip()
                    if len(ex_str) > 50:
                        ex_str = ex_str[:47] + "..."
                    cleaned_examples.append(ex_str)
                examples_str = "; ".join(cleaned_examples)
            else:
                examples_str = f"{invalid_count} issues found (examples not captured)"
            
            validation_results.append({
                'Column': column,
                'Dimension': rule.get('Dimension'),
                'Source': rule.get('Source', 'Unknown'),
                'Invalid_Count': int(invalid_count),
                'Issues_Found_Example': examples_str
            })
        
        return pd.DataFrame(validation_results)


# ==========================================
# ENHANCED VALIDATION DETECTION
# ==========================================

class DynamicValidationDetector:
    """Enhanced validation detection combining AI and statistical analysis"""
    
    def __init__(self):
        self.ai_engine = AIValidationEngine()
        self.excel_notes = {}  # Store Excel notes/comments
    
    def detect_all_validations(self, df: pd.DataFrame, profiles: Dict, uploaded_file=None, sheet_name: str = None) -> List[Dict]:
            """Detect all validations using AI-powered analysis with robust error handling

            IMPORTANT: Generates rules for ALL columns regardless of null percentage
            
            Args:
                df: DataFrame to analyze
                profiles: Column profiles dictionary
                uploaded_file: Original uploaded file (for Excel metadata extraction)
                sheet_name: Sheet name (for Excel files)
            """
            logger.info("="*80)
            logger.info("STARTING VALIDATION RULE GENERATION")
            logger.info(f"Total columns in dataframe: {len(df.columns)}")
            logger.info(f"Total profiles available: {len(profiles)}")
            logger.info("IMPORTANT: Processing ALL columns regardless of null percentage")
            logger.info("="*80)
            
            # Extract Excel notes/comments if available
            if uploaded_file is not None:
                try:
                    logger.info("Attempting to extract Excel metadata and notes...")
                    self.excel_notes = extract_excel_cell_notes(uploaded_file, sheet_name)
                    if self.excel_notes:
                        logger.info(f"Successfully extracted notes for {len(self.excel_notes)} columns")
                    else:
                        logger.info("No Excel notes found or file is not Excel format")
                except Exception as e:
                    logger.warning(f"Could not extract Excel notes: {str(e)}")
                    self.excel_notes = {}

            all_validations = []

            progress_bar = st.progress(0)
            status_text = st.empty()
            total_cols = len(df.columns)

            successful_cols = 0
            failed_cols = []
            skipped_cols = 0

            # Add rate limiting counter
            api_calls = 0
            max_api_calls_per_minute = AzureOpenAIConfig.MAX_REQUESTS_PER_MINUTE

            logger.info(f"Rate limit configuration: {max_api_calls_per_minute} requests/minute")

            for idx, col in enumerate(df.columns):
                try:
                    logger.info(f"\n--- Processing column {idx + 1}/{total_cols}: {col} ---")

                    progress_bar.progress((idx + 1) / total_cols)
                    status_text.text(f"Analyzing column {idx + 1}/{total_cols}: {col} ({successful_cols} successful, {len(failed_cols)} failed)")

                    if col not in profiles:
                        logger.warning(f"Column '{col}' not found in profiles. Skipping.")
                        skipped_cols += 1
                        continue

                    profile = profiles[col]
                    logger.info(f"Column '{col}' - Type: {profile.dtype}, Null%: {profile.null_percentage:.1f}%")

                    # REMOVED: Skip columns with too many nulls
                    # We now process ALL columns regardless of null percentage
                    # if profile.null_percentage > 90:
                    #     logger.info(f"Column '{col}' has {profile.null_percentage:.1f}% nulls. Skipping.")
                    #     skipped_cols += 1
                    #     continue

                    # Rate limiting - pause if needed
                    if api_calls >= max_api_calls_per_minute:
                        logger.warning(f"Rate limit reached ({api_calls} calls). Pausing for 60 seconds...")
                        status_text.text(f"Rate limit reached. Pausing for 60 seconds... ({successful_cols}/{total_cols} completed)")
                        time.sleep(60)
                        api_calls = 0
                        logger.info("Resuming after rate limit pause")

                    # Generate AI-powered rules with retry logic
                    max_retries = 2
                    rules_generated = False

                    for attempt in range(max_retries):
                        try:
                            logger.info(f"Attempt {attempt + 1}/{max_retries} to generate rules for '{col}'")
                            rules = self.ai_engine.generate_dynamic_rules(df, col, profile, excel_notes=self.excel_notes)

                            if rules:  # Only count if rules were actually generated
                                logger.info(f"SUCCESS: Generated {len(rules)} rules for '{col}'")
                                for rule in rules:
                                    logger.debug(f"  - {rule.get('Dimension')}: {rule.get('Data Quality Rule')}")

                                all_validations.extend(rules)
                                successful_cols += 1
                                api_calls += 1
                                rules_generated = True
                                break
                            else:
                                logger.warning(f"No rules returned for '{col}' on attempt {attempt + 1}")

                        except Exception as e:
                            error_msg = str(e)
                            logger.error(f"ERROR on attempt {attempt + 1} for '{col}': {error_msg}")

                            if "rate" in error_msg.lower() or "quota" in error_msg.lower():
                                # Rate limit hit - wait longer
                                logger.warning("Rate limit error detected. Waiting 60 seconds...")
                                status_text.text(f"Rate limit hit. Waiting 60 seconds...")
                                time.sleep(60)
                                api_calls = 0
                            elif attempt == max_retries - 1:
                                # Last attempt failed
                                logger.error(f"FAILED: All attempts failed for '{col}'. Using fallback rules.")
                                failed_cols.append(col)
                                # Generate basic fallback rules
                                fallback_rules = self._generate_fallback_rules(col, profile)
                                logger.info(f"Generated {len(fallback_rules)} fallback rules for '{col}'")
                                all_validations.extend(fallback_rules)
                            else:
                                # Wait before retry
                                logger.info("Waiting 2 seconds before retry...")
                                time.sleep(2)

                    # Small delay between API calls to avoid rate limits
                    if rules_generated:
                        time.sleep(0.5)  # 500ms delay between successful calls

                except Exception as e:
                    logger.error(f"EXCEPTION processing column '{col}': {str(e)}", exc_info=True)
                    failed_cols.append(col)
                    # Generate basic fallback rules
                    try:
                        fallback_rules = self._generate_fallback_rules(col, profile)
                        logger.info(f"Generated {len(fallback_rules)} fallback rules for '{col}' after exception")
                        all_validations.extend(fallback_rules)
                    except Exception as fallback_error:
                        logger.error(f"Failed to generate fallback rules for '{col}': {str(fallback_error)}")

            progress_bar.empty()
            status_text.empty()

            # Log final summary
            logger.info("="*80)
            logger.info("VALIDATION RULE GENERATION COMPLETE")
            logger.info(f"Total columns processed: {total_cols}")
            logger.info(f"Successfully generated (AI): {successful_cols}")
            logger.info(f"Fallback rules used: {len(failed_cols)}")
            logger.info(f"Skipped (missing profile): {skipped_cols}")
            logger.info(f"Total rules generated: {len(all_validations)}")
            logger.info(f"Failed columns: {', '.join(failed_cols) if failed_cols else 'None'}")
            logger.info("="*80)

            # Show comprehensive summary
            st.success(f"✅ **Validation Rules Generated:**")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Columns", total_cols)
            with col2:
                st.metric("AI Generated", successful_cols, delta="Success")
            with col3:
                st.metric("Fallback Rules", len(failed_cols), delta="Used fallback")
            with col4:
                st.metric("Skipped", skipped_cols, delta="Missing profile")

            if failed_cols:
                with st.expander(f"ℹ️ Columns using fallback rules ({len(failed_cols)})"):
                    st.write(", ".join(failed_cols[:20]))
                    if len(failed_cols) > 20:
                        st.caption(f"... and {len(failed_cols) - 20} more")

            # Show log file location
            st.info(f"📝 Detailed logs saved to: data_profiler_validation.log")

            return all_validations

    
    def _generate_fallback_rules(self, column_name: str, profile: Any) -> List[Dict]:
        """Generate basic validation rules when AI fails"""
        rules = []
        
        # Completeness rule
        if profile.null_percentage > 0:
            rules.append({
                'S.No': 1,
                'Column': column_name,
                'Business Field': column_name,
                'Dimension': 'Completeness',
                'Data Quality Rule': f'{column_name} should not be blank',
                'Source': 'AI Generated'
            })
        
        # Uniqueness rule for high unique percentage
        if profile.unique_percentage > 95:
            rules.append({
                'S.No': len(rules) + 1,
                'Column': column_name,
                'Business Field': column_name,
                'Dimension': 'Uniqueness',
                'Data Quality Rule': f'{column_name} must be unique',
                'Source': 'AI Generated'
            })
        
        # Validity rule based on data type
        dtype_lower = str(profile.dtype).lower()
        if 'int' in dtype_lower or 'float' in dtype_lower:
            rules.append({
                'S.No': len(rules) + 1,
                'Column': column_name,
                'Business Field': column_name,
                'Dimension': 'Validity',
                'Data Quality Rule': f'{column_name} must be numeric',
                'Source': 'AI Generated'
            })
        elif 'date' in dtype_lower:
            rules.append({
                'S.No': len(rules) + 1,
                'Column': column_name,
                'Business Field': column_name,
                'Dimension': 'Validity',
                'Data Quality Rule': f'{column_name} must be a valid date',
                'Source': 'AI Generated'
            })
        
        return rules
    
    def generate_comprehensive_dq_rules(self, df: pd.DataFrame, profiles: Dict, uploaded_file=None, sheet_name: str = None) -> pd.DataFrame:
        """Generate comprehensive DQ rules with validation results
        
        Args:
            df: DataFrame to analyze
            profiles: Column profiles dictionary
            uploaded_file: Original uploaded file (for Excel metadata extraction)
            sheet_name: Sheet name (for Excel files)
        """
        logger.info("="*80)
        logger.info("STARTING COMPREHENSIVE DQ RULES GENERATION")
        logger.info(f"DataFrame shape: {df.shape}")
        logger.info(f"Profiles count: {len(profiles)}")
        logger.info("="*80)
        
        all_rules = self.detect_all_validations(df, profiles, uploaded_file=uploaded_file, sheet_name=sheet_name)
        
        logger.info(f"\nTotal rules collected from detect_all_validations: {len(all_rules)}")
        
        # Log unique columns in rules
        unique_cols_in_rules = set(rule.get('Column') for rule in all_rules)
        logger.info(f"Unique columns with rules: {len(unique_cols_in_rules)}")
        logger.info(f"Columns: {', '.join(sorted(unique_cols_in_rules)[:10])}...")

        # Validate data against rules
        logger.info("\nValidating data against rules...")
        validation_results = self.ai_engine.validate_data_against_rules(df, all_rules)
        logger.info(f"Validation results shape: {validation_results.shape if not validation_results.empty else 'Empty'}")

        # Aggregate validation results by (Column, Dimension, Source)
        if not validation_results.empty:
            logger.info("Aggregating validation results...")
            validation_results = validation_results.copy()
            validation_results['Column'] = validation_results['Column'].astype(str)
            validation_results['Dimension'] = validation_results['Dimension'].astype(str)
            validation_results['Source'] = validation_results['Source'].astype(str)
            
            agg_val = validation_results.groupby(['Column', 'Dimension', 'Source'], sort=False).agg({
                'Invalid_Count': 'sum',
                'Issues_Found_Example': lambda x: '; '.join([str(v) for v in x if pd.notna(v) and v != "✓ All values valid - No issues found"]) or "✓ All values valid - No issues found"
            }).reset_index()
            logger.info(f"Aggregated validation results: {len(agg_val)} rows")
        else:
            logger.warning("No validation results to aggregate")
            agg_val = pd.DataFrame(columns=['Column', 'Dimension', 'Source', 'Invalid_Count', 'Issues_Found_Example'])

        # Merge rules having the same (Column, Dimension, Source)
        logger.info("\nMerging rules by (Column, Dimension, Source)...")
        rules_map = {}
        for rule in all_rules:
            col = rule.get('Column')
            dim = rule.get('Dimension')
            src = rule.get('Source', 'Unknown')
            key = (col, dim, src)
            
            rules_map.setdefault(key, {
                'Business Field': rule.get('Business Field'),
                'Rules': []
            })
            
            stmt = rule.get('Data Quality Rule')
            if stmt and stmt not in rules_map[key]['Rules']:
                rules_map[key]['Rules'].append(stmt)
        
        logger.info(f"Unique (Column, Dimension) combinations: {len(rules_map)}")

        # Build output rows from aggregated map
        logger.info("\nBuilding output dataframe...")
        output_data = []
        for idx, ((col, dim, src), meta) in enumerate(rules_map.items(), 1):
            match = agg_val[(agg_val['Column'] == col) & 
                           (agg_val['Dimension'] == dim) & 
                           (agg_val['Source'] == src)]
            
            invalid_count = int(match.iloc[0]['Invalid_Count']) if not match.empty else 0
            issues_example = match.iloc[0]['Issues_Found_Example'] if not match.empty else "✓ All values valid - No issues found"

            # Determine source label
            if src == 'Client Extracted':
                source_label = '✓ Client provided rule - From Excel metadata'
            elif src == 'AI Generated':
                source_label = '🤖 AI Generated'
            else:
                source_label = src

            row = {
                'S.No': idx,
                'Column': col,
                'Business Field': meta.get('Business Field'),
                'Dimension': dim,
                'Data Quality Rule': '; '.join(meta.get('Rules', [])),
                'Issues Found': invalid_count,
                'Issues Found Example': issues_example,
                'Source': source_label
            }
            output_data.append(row)
        
        result_df = pd.DataFrame(output_data)
        
        logger.info("="*80)
        logger.info("COMPREHENSIVE DQ RULES GENERATION COMPLETE")
        logger.info(f"Final dataframe shape: {result_df.shape}")
        logger.info(f"Total rules in output: {len(result_df)}")
        logger.info(f"Unique columns in output: {result_df['Column'].nunique() if not result_df.empty else 0}")
        logger.info("="*80)

        return result_df


# ==========================================
# CLIENT RULE EXTRACTION FROM EXCEL METADATA
# ==========================================

def _extract_client_rules_from_excel_metadata() -> List[Dict]:
    """
    Extract client-provided data quality rules from Excel file metadata.
    
    Reads the raw Excel file to detect header patterns like:
    Row 1: Descriptive instructions
    Row 2: Data type constraints (Number, 30 Characters, YYYY/MM/DD, etc.)
    Row 3: Column headers (with * for mandatory fields)
    
    Creates human-readable rules like:
    - "Asset Description should be maximum 80 characters"
    - "Interface Line Number must be Number"
    - "Date Placed in Service must be in YYYY/MM/DD date format"
    """
    try:
        state = st.session_state.app_state
        config = st.session_state.get('excel_config', {})
        
        # Get file path and sheet name
        file_path = getattr(state, 'file_path', None) or config.get('file_path')
        sheet_name = getattr(state, 'sheet_name', None) or config.get('selected_sheet')
        
        if not file_path or not os.path.exists(file_path):
            return []
        
        if not sheet_name:
            return []
        
        # Read raw Excel with first 4 rows to get metadata
        df_meta = pd.read_excel(file_path, sheet_name=sheet_name, header=None, nrows=4)
        
        # Row 2 contains data type constraints
        # Row 3 contains column headers
        if len(df_meta) < 4:
            return []
        
        constraints_row = df_meta.iloc[2]  # Row index 2 (3rd row)
        headers_row = df_meta.iloc[3]      # Row index 3 (4th row)
        
        extracted_rules = []
        
        # Iterate through columns
        for col_idx in range(len(constraints_row)):
            constraint = constraints_row.iloc[col_idx]
            header = headers_row.iloc[col_idx]
            
            # Skip if either is null/empty
            if pd.isna(constraint) or pd.isna(header):
                continue
            
            constraint_str = str(constraint).strip()
            header_str = str(header).strip()
            
            # Skip empty strings
            if not constraint_str or not header_str or constraint_str.lower() == 'nan' or header_str.lower() == 'nan':
                continue
            
            # Clean column name (remove * prefix for mandatory indicator)
            clean_column_name = header_str.lstrip('*').strip()
            is_mandatory = header_str.startswith('*')
            
            # Parse different constraint types
            rules = _parse_constraint_to_rules(constraint_str, clean_column_name, is_mandatory)
            
            # Add to extracted rules
            for rule in rules:
                # Try to find matching column in actual dataframe
                actual_col = _find_matching_column(clean_column_name, state.df)
                
                extracted_rules.append({
                    'S.No': len(extracted_rules) + 1,
                    'Column': actual_col if actual_col else clean_column_name,
                    'Business Field': clean_column_name,
                    'Dimension': rule['dimension'],
                    'Data Quality Rule': rule['rule_statement'],
                    'Issues Found': 0,
                    'Issues Found Example': '✓ Client provided rule - From Excel metadata',
                    'Source': 'Client Extracted',  # Mark source
                    'Metadata_Row': 2
                })
        
        return extracted_rules
        
    except Exception as e:
        st.warning(f"Could not extract client rules from Excel metadata: {str(e)}")
        return []


def _parse_constraint_to_rules(constraint: str, column_name: str, is_mandatory: bool) -> List[Dict]:
    """
    Parse a constraint string into human-readable validation rules.
    
    Examples:
    - "30 Characters" -> "Column should be maximum 30 characters"
    - "Number" -> "Column must be numeric"
    - "YYYY/MM/DD" -> "Column must be in YYYY/MM/DD date format"
    - "Number without thousand separator" -> Multiple rules
    """
    rules = []
    constraint_lower = constraint.lower().strip()
    
    # Pattern 1: Character length constraints (e.g., "30 Characters", "80 Characters")
    char_match = re.match(r'^(\d+)\s*characters?$', constraint_lower)
    if char_match:
        max_chars = char_match.group(1)
        rules.append({
            'dimension': 'Character Length',
            'rule_statement': f"{column_name} should be maximum {max_chars} characters"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
        return rules
    
    # Pattern 2: Date format constraints (e.g., "YYYY/MM/DD")
    if 'yyyy' in constraint_lower or 'mm' in constraint_lower or 'dd' in constraint_lower:
        rules.append({
            'dimension': 'Validity',
            'rule_statement': f"{column_name} must be in {constraint} date format"
        })
        rules.append({
            'dimension': 'Timeliness',
            'rule_statement': f"{column_name} should not be future dated"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
        return rules
    
    # Pattern 3: Number constraints
    if constraint_lower == 'number':
        rules.append({
            'dimension': 'Validity',
            'rule_statement': f"{column_name} must be numeric"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
        return rules
    
    # Pattern 4: Number without thousand separator
    if 'number without thousand separator' in constraint_lower:
        rules.append({
            'dimension': 'Validity',
            'rule_statement': f"{column_name} must be numeric"
        })
        rules.append({
            'dimension': 'Conformity',
            'rule_statement': f"{column_name} must not contain thousand separators (commas)"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
        return rules
    
    # Pattern 5: Generic text constraint
    if 'text' in constraint_lower or 'string' in constraint_lower:
        rules.append({
            'dimension': 'Validity',
            'rule_statement': f"{column_name} must contain valid text"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
        return rules
    
    # Default: Treat as conformity rule
    if rules == []:
        rules.append({
            'dimension': 'Conformity',
            'rule_statement': f"{column_name} must conform to format: {constraint}"
        })
        if is_mandatory:
            rules.append({
                'dimension': 'Completeness',
                'rule_statement': f"{column_name} must not be blank"
            })
    
    return rules


def _find_matching_column(header_name: str, df: pd.DataFrame) -> Optional[str]:
    """
    Find the actual column in dataframe that matches the header name.
    Handles cases where user selected a different header row.
    """
    if df is None:
        return None
    
    header_lower = header_name.lower().strip()
    
    # Direct match
    for col in df.columns:
        if col.lower().strip() == header_lower:
            return col
    
    # Partial match
    for col in df.columns:
        if header_lower in col.lower() or col.lower() in header_lower:
            return col
    
    return None


# ==========================================
# CORE FUNCTIONS
# ==========================================

def generate_match_rules(df, profiles):
    """Generate match rules - excludes 100% unique columns from exact match"""
    rules = []
    counter = 1

    # Analyze columns
    analysis = {}
    for col, prof in profiles.items():
        total_rows = int(getattr(prof, 'total_rows', 0) or 0)
        unique_count = int(getattr(prof, 'unique_count', 0) or 0)
        dup_count = max(0, total_rows - unique_count)
        dup_pct = (dup_count / total_rows * 100) if total_rows > 0 else 0

        dtype = getattr(prof, 'dtype', '') or ''

        analysis[col] = {
            'null_pct': getattr(prof, 'null_percentage', 0),
            'unique_pct': getattr(prof, 'unique_percentage', 0),
            'dup_pct': dup_pct,
            'dup_count': dup_count,
            'is_text': dtype == 'object',
            'is_num': any(t in dtype for t in ['int', 'float']),
            'avg_len': getattr(prof, 'avg_length', 0),
            'max_len': getattr(prof, 'max_length', 0),
            'min_len': getattr(prof, 'min_length', 0),
            'total_rows': total_rows
        }

    # EXACT MATCH CANDIDATES
    exact_candidates = []
    for col, a in analysis.items():
        if a['unique_pct'] == 100 or a['dup_count'] == 0:
            continue

        score = 0
        reasons = []

        if 95 <= a['unique_pct'] < 100:
            score += 35
            reasons.append(f"Near-unique ({a['unique_pct']:.1f}%)")
        elif 80 <= a['unique_pct'] < 95:
            score += 25
            reasons.append(f"High uniqueness ({a['unique_pct']:.1f}%)")

        if a['null_pct'] < 1:
            score += 20
            reasons.append("Complete data (no nulls)")
        elif a['null_pct'] < 5:
            score += 15
            reasons.append("Low null rate")

        if a['is_text'] and a['max_len'] == a['min_len'] and 4 <= a['avg_len'] <= 20:
            score += 25
            reasons.append(f"Fixed length ({int(a['avg_len'])} chars)")

        if 0 < a['dup_pct'] < 5:
            score += 15
            reasons.append(f"Low duplicates ({a['dup_pct']:.1f}%)")

        if score >= 50 and a['dup_count'] > 0:
            exact_candidates.append({
                'column': col,
                'score': score,
                'reasons': reasons,
                'dup_count': a['dup_count'],
                'unique_pct': a['unique_pct']
            })

    exact_candidates.sort(key=lambda x: x['score'], reverse=True)

    for cand in exact_candidates[:4]:
        prob = "Strongest" if cand['score'] >= 85 else "Very Strong" if cand['score'] >= 75 else "Strong" if cand['score'] >= 65 else "Good"
        rules.append({
            'Rule No': f"R{counter:02d}",
            'Rule Type': 'Exact',
            'Columns': cand['column'],
            'Match Probability': prob,
            'Rationale': f"{'; '.join(cand['reasons'][:2])} | Duplicates: {cand['dup_count']}",
            'Confidence': cand['score']
        })
        counter += 1

    # FUZZY MATCH CANDIDATES
    fuzzy_candidates = []
    for col, a in analysis.items():
        if not a['is_text']:
            continue

        score = 0
        reasons = []

        if 30 <= a['unique_pct'] <= 90:
            score += 35
            reasons.append(f"Medium uniqueness ({a['unique_pct']:.1f}%)")
        elif 10 <= a['unique_pct'] < 30:
            score += 20
            reasons.append(f"Low-medium uniqueness ({a['unique_pct']:.1f}%)")

        if 10 <= a['avg_len'] <= 100:
            score += 25
            reasons.append(f"Name/description length ({a['avg_len']:.0f} chars)")
        elif a['avg_len'] > 100:
            score += 15
            reasons.append("Long text field")

        name_indicators = ['name', 'desc', 'title', 'product', 'customer', 'company', 'vendor', 'supplier', 'brand', 'item']
        if any(ind in col.lower() for ind in name_indicators):
            score += 25
            reasons.append("Name/description column")

        if a['dup_count'] > 1:
            score += 15
            reasons.append(f"Has duplicates to match ({a['dup_count']})")

        if score >= 45:
            fuzzy_candidates.append({
                'column': col,
                'score': score,
                'reasons': reasons,
                'unique_pct': a['unique_pct']
            })

    fuzzy_candidates.sort(key=lambda x: x['score'], reverse=True)

    for cand in fuzzy_candidates[:4]:
        prob = "Strong" if cand['score'] >= 70 else "Good" if cand['score'] >= 60 else "Medium"
        rules.append({
            'Rule No': f"R{counter:02d}",
            'Rule Type': 'Fuzzy',
            'Columns': cand['column'],
            'Match Probability': prob,
            'Rationale': '; '.join(cand['reasons'][:3]),
            'Confidence': cand['score']
        })
        counter += 1

    # COMBINED RULES
    if exact_candidates and fuzzy_candidates:
        for e in exact_candidates[:2]:
            for f in fuzzy_candidates[:3]:
                if e['column'] != f['column'] and len(rules) < 10:
                    score = (e['score'] + f['score']) / 2
                    prob = "Enterprise" if score >= 75 else "Strong" if score >= 65 else "Good"
                    rules.append({
                        'Rule No': f"R{counter:02d}",
                        'Rule Type': 'Combined',
                        'Columns': f"{e['column']} (Exact) + {f['column']} (Fuzzy)",
                        'Match Probability': prob,
                        'Rationale': f"Exact on {e['column']}; Fuzzy match on {f['column']}",
                        'Confidence': score
                    })
                    counter += 1
                    if len(rules) >= 10:
                        break
            if len(rules) >= 10:
                break

    if not rules and analysis:
        first = list(analysis.keys())[0]
        rules.append({
            'Rule No': 'R01',
            'Rule Type': 'Exact',
            'Columns': first,
            'Match Probability': 'Medium',
            'Rationale': 'Default fallback rule',
            'Confidence': 50
        })

    rules.sort(key=lambda x: x['Confidence'], reverse=True)
    for i, r in enumerate(rules[:10], 1):
        r['Rule No'] = f"R{i:02d}"

    return rules[:10]


def find_duplicate_groups(df, col):
    """Find duplicate groups in a column"""
    if col not in df.columns:
        return []

    value_counts = df[col].value_counts()
    duplicates = value_counts[value_counts > 1]

    groups = []
    for val, count in duplicates.head(10).items():
        matching_rows = df[df[col] == val]
        groups.append({
            'value': str(val)[:50],
            'count': int(count),
            'percentage': round((count / len(df)) * 100, 2),
            'row_indices': matching_rows.index.tolist()[:5]
        })

    return groups


def safe_get_special_chars(prof):
    try:
        if hasattr(prof, 'special_chars') and prof.special_chars:
            return [c for c in prof.special_chars if isinstance(c, dict) and 'count' in c]
    except: 
        pass
    return []


def _remove_ghost_columns():
    state = st.session_state.app_state
    if state.df is None: 
        return False
    ghosts = [c for c in state.df.columns if str(c).startswith('Unnamed:') and state.df[c].isnull().all()]
    if ghosts:
        state.df = state.df.drop(columns=ghosts)
        for c in ghosts: 
            state.column_profiles.pop(c, None)
        show_toast(f"Removed {len(ghosts)} empty columns", "success")
        return True
    return False


# ==========================================
# UI RENDERING FUNCTIONS
# ==========================================

def render_data_profiling():
    state = st.session_state.app_state
    if state.df is None:
        st.info("📤 Please load data first")
        return

    _remove_ghost_columns()

    _render_executive_dashboard()

    tabs = st.tabs([
        "📊 Overview", 
        "📋 Column Profiles", 
        "🎯 Match Rules", 
        "📥 Export"
    ])

    with tabs[0]: 
        _render_overview_tab()
    with tabs[1]: 
        _render_profiles_tab()
    with tabs[2]: 
        _render_match_rules_tab()
    with tabs[3]: 
        _render_export_tab()


def _render_executive_dashboard():
    state = st.session_state.app_state
    df = state.df
    profiles = state.column_profiles

    total_rows = len(df)
    total_cols = len(df.columns)
    total_cells = total_rows * total_cols
    missing_cells = sum(p.null_count for p in profiles.values())
    completeness = ((total_cells - missing_cells) / total_cells) * 100 if total_cells else 0

    quality_scores = []
    for p in profiles.values():
        comp = getattr(p, 'non_null_percentage', 100)
        uniq = min(100, p.unique_percentage * 1.2) if p.unique_percentage < 100 else 90
        consistency = 100 - (20 if hasattr(p, 'formatting_info') and p.formatting_info and not p.formatting_info.get('consistent_case', True) else 0)
        validity = 100 - min(20, len(safe_get_special_chars(p)) * 2)
        quality_scores.append(comp * 0.4 + uniq * 0.3 + consistency * 0.2 + validity * 0.1)
    avg_quality = round(sum(quality_scores)/len(quality_scores), 1) if quality_scores else 0

    cols = st.columns(6)
    kpi_data = [
        ("📊", "Rows", f"{total_rows:,}"),
        ("📁", "Columns", total_cols),
        ("⭐", "Quality", f"{avg_quality:.0f}%"),
        ("✅", "Completeness", f"{completeness:.1f}%"),
        ("⚠️", "Missing", f"{missing_cells:,}"),
        ("🎯", "Fill Rate", f"{100-(missing_cells/total_cells*100):.1f}%" if total_cells else "N/A")
    ]

    for col, (icon, label, value) in zip(cols, kpi_data):
        with col:
            st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 15px; border-radius: 10px; text-align: center; color: white;">
                <div style="font-size: 24px;">{icon}</div>
                <div style="font-size: 12px; opacity: 0.9;">{label}</div>
                <div style="font-size: 20px; font-weight: bold;">{value}</div>
            </div>
            """, unsafe_allow_html=True)

    st.divider()


def _render_overview_tab():
    state = st.session_state.app_state
    profiles = state.column_profiles
    df = state.df

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Data Type Distribution")
        dtype_counts = {}
        for p in profiles.values():
            t = 'Numeric' if any(x in p.dtype for x in ['int', 'float']) else 'DateTime' if 'date' in p.dtype else 'Text' if p.dtype == 'object' else 'Other'
            dtype_counts[t] = dtype_counts.get(t, 0) + 1

        fig = px.pie(values=list(dtype_counts.values()), names=list(dtype_counts.keys()),
                    hole=0.4, color_discrete_sequence=px.colors.qualitative.Set3)
        fig.update_traces(textposition='inside', textinfo='percent+label')
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.subheader("Data Quality by Column")
        quality_data = []
        for col, p in profiles.items():
            comp = getattr(p, 'non_null_percentage', 100)
            uniq = p.unique_percentage
            quality_data.append({'Column': col, 'Completeness': comp, 'Uniqueness': uniq})

        if quality_data:
            qdf = pd.DataFrame(quality_data).sort_values('Completeness').head(15)
            fig = px.bar(qdf, x='Column', y=['Completeness', 'Uniqueness'],
                        barmode='group', color_discrete_sequence=['#10b981', '#3b82f6'])
            fig.update_layout(xaxis_tickangle=-45, yaxis_title="Percentage")
            st.plotly_chart(fig, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Missing Data by Column")
        null_data = [(c, p.null_percentage) for c, p in profiles.items() if p.null_percentage > 0]
        if null_data:
            null_df = pd.DataFrame(null_data, columns=['Column', 'Null %']).sort_values('Null %', ascending=True).tail(10)
            fig = px.bar(null_df, x='Null %', y='Column', orientation='h', color='Null %',
                        color_continuous_scale='Reds')
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.success("✅ No missing data!")

    with col2:
        st.subheader("Column Cardinality")
        card_data = [{'Column': c, 'Unique Values': p.unique_count, 'Type': 'Numeric' if any(x in p.dtype for x in ['int', 'float']) else 'Other'}
                     for c, p in profiles.items()]
        
        if card_data:
            card_df = pd.DataFrame(card_data).sort_values('Unique Values', ascending=False).head(15)
            fig = px.scatter(card_df, x='Column', y='Unique Values', color='Type', size='Unique Values',
                            color_discrete_sequence=px.colors.qualitative.Bold)
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No cardinality data available")


def _render_profiles_tab():
    state = st.session_state.app_state
    profiles = state.column_profiles
    df = state.df

    c1, c2, c3 = st.columns([3, 1, 1])
    with c1: 
        search = st.text_input("🔍 Search columns")
    with c2: 
        dtype_filter = st.selectbox("Type", ["All", "Numeric", "Text", "Date", "High Risk"])
    with c3: 
        sort_by = st.selectbox("Sort", ["Name", "Null %", "Uniqueness", "Risk"])

    filtered = {k: v for k, v in profiles.items() if not search or search.lower() in k.lower()}

    if dtype_filter == "Numeric":
        filtered = {k: v for k, v in filtered.items() if any(t in v.dtype for t in ['int', 'float'])}
    elif dtype_filter == "Text":
        filtered = {k: v for k, v in filtered.items() if v.dtype == 'object'}
    elif dtype_filter == "Date":
        filtered = {k: v for k, v in filtered.items() if 'date' in v.dtype}
    elif dtype_filter == "High Risk":
        filtered = {k: v for k, v in filtered.items() if getattr(v, 'risk_level', 'Low') == 'High'}

    if sort_by == "Null %":
        filtered = dict(sorted(filtered.items(), key=lambda x: x[1].null_percentage, reverse=True))
    elif sort_by == "Uniqueness":
        filtered = dict(sorted(filtered.items(), key=lambda x: x[1].unique_percentage, reverse=True))
    elif sort_by == "Risk":
        filtered = dict(sorted(filtered.items(), key=lambda x: getattr(x[1], 'risk_score', 0), reverse=True))

    st.write(f"Showing {len(filtered)} of {len(profiles)} columns")

    cols_per_row = 2
    items = list(filtered.items())  # Remove [:30] limit - show all columns

    for i in range(0, len(items), cols_per_row):
        cols = st.columns(cols_per_row)

        for col_ui, (col, prof) in zip(cols, items[i:i+cols_per_row]):
            with col_ui:
                with st.expander(
                    f"📊 {col} | {prof.dtype} | Quality: {100-prof.null_percentage:.0f}%"
                ):
                    c1, c2, c3, c4 = st.columns(4)

                    with c1:
                        st.markdown("**Volume**")
                        st.write(f"Rows: {prof.total_rows:,}")
                        st.write(f"Non-null: {getattr(prof, 'non_null_count', prof.total_rows - prof.null_count):,}")
                        st.write(f"Null: {prof.null_count:,} ({prof.null_percentage:.1f}%)")

                    with c2:
                        st.markdown("**Uniqueness**")
                        dup = prof.total_rows - prof.unique_count
                        st.write(f"Unique: {prof.unique_count:,}")
                        st.write(f"Duplicates: {dup:,}")
                        st.write(f"Unique %: {prof.unique_percentage:.1f}%")

                    with c3:
                        st.markdown("**Length**")
                        st.write(f"Min: {getattr(prof, 'min_length', 'N/A')}")
                        st.write(f"Max: {getattr(prof, 'max_length', 'N/A')}")
                        st.write(f"Avg: {getattr(prof, 'avg_length', 0):.1f}")

                    with c4:
                        st.markdown("**Risk**")
                        risk = getattr(prof, 'risk_level', 'Low')
                        color = "🔴" if risk == "High" else "🟡" if risk == "Medium" else "🟢"
                        st.write(f"Level: {color} {risk}")
                        st.write(f"Score: {getattr(prof, 'risk_score', 0)}/100")

                    if prof.unique_count < prof.total_rows:
                        dups = find_duplicate_groups(df, col)
                        if dups:
                            with st.expander(f"⚠️ Duplicates ({len(dups)} groups)"):
                                for d in dups[:5]:
                                    st.write(
                                        f"- Value '{d['value'][:30]}' "
                                        f"appears {d['count']} times ({d['percentage']}%)"
                                    )


def _render_ai_validations_tab():
    """UNIFIED AI Validations Tab - Single Combined Table"""
    state = st.session_state.app_state
    df = state.df
    profiles = state.column_profiles
    
    missing_config = AzureOpenAIConfig.validate()
    if missing_config:
        st.error(f"⚠️ Azure OpenAI not configured. Missing: {', '.join(missing_config)}")
        st.info("Please set these in your Streamlit secrets or environment variables.")
        return
    
    # Initialize state for unified rules
    if not hasattr(state, 'unified_validation_rules'):
        state.unified_validation_rules = None
        state.unified_rules_generated = False
    
    st.markdown("""
    <div style="margin-bottom: 15px;">
        <h4 style="margin: 0;">🎯 Complete Data Quality Analysis</h4>
        <p style="color: #666; margin: 5px 0 0 0; font-size: 13px;">
            Click the button below to perform 100% complete analysis:
            <br>✓ Extract client rules from Excel metadata
            <br>✓ Generate AI-powered rules for all columns
            <br>✓ Display everything in a single unified table
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns([1, 3])
    with col1:
        if not state.unified_rules_generated:
            if st.button("🚀 Generate Complete Analysis", type="primary"):
                with st.spinner("🔍 Performing 100% complete analysis..."):
                    # Step 1: Extract client rules from Excel metadata
                    client_rules = _extract_client_rules_from_excel_metadata()
                    
                    # Step 2: Generate comprehensive rules (client + AI combined)
                    detector = DynamicValidationDetector()
                    unified_df = detector.generate_comprehensive_dq_rules(df, profiles, client_rules)
                    
                    # Store in state for persistence
                    state.unified_validation_rules = unified_df
                    state.unified_rules_generated = True
                    
                    from state.session import _save_persisted_data
                    _save_persisted_data()
                    
                    st.success(f"✅ Complete! Generated {len(unified_df)} total rules (Client + AI)")
                    st.rerun()
        else:
            if st.button("🗑️ Clear All Rules", type="secondary"):
                state.unified_rules_generated = False
                state.unified_validation_rules = pd.DataFrame()
                
                from state.session import _save_persisted_data
                _save_persisted_data()
                st.rerun()
    
    # Display unified table if generated
    if state.unified_rules_generated and state.unified_validation_rules is not None and not state.unified_validation_rules.empty:
        validation_df = state.unified_validation_rules
        
        # Summary stats
        total_rules = len(validation_df)
        client_rules_count = len(validation_df[validation_df['Source'].str.contains('Client', na=False)])
        ai_rules_count = len(validation_df[validation_df['Source'].str.contains('AI', na=False)])
        
        st.markdown(f"""
        <div style="background-color: #f0f9ff; border-left: 4px solid #3b82f6; padding: 15px; margin-bottom: 20px;">
            <strong>📊 Analysis Summary:</strong><br>
            Total Rules: {total_rules} | Client Rules: {client_rules_count} ✓ | AI Rules: {ai_rules_count} 🤖
        </div>
        """, unsafe_allow_html=True)
        
        # Filter by Dimension
        available_dims = validation_df['Dimension'].unique().tolist()
        selected_dims = st.multiselect(
            "Filter by DQ Dimension", 
            available_dims, 
            default=available_dims,
            key="unified_dimension_filter"
        )
        
        if selected_dims:
            filtered_df = validation_df[validation_df['Dimension'].isin(selected_dims)]
        else:
            filtered_df = validation_df
        
        # Filter by Source
        col1, col2 = st.columns(2)
        with col1:
            show_client = st.checkbox("Show Client Rules", value=True, key="show_client")
        with col2:
            show_ai = st.checkbox("Show AI Rules", value=True, key="show_ai")
        
        if show_client and not show_ai:
            filtered_df = filtered_df[filtered_df['Source'].str.contains('Client', na=False)]
        elif show_ai and not show_client:
            filtered_df = filtered_df[filtered_df['Source'].str.contains('AI', na=False)]
        elif not show_client and not show_ai:
            filtered_df = pd.DataFrame()  # Empty
        
        if not filtered_df.empty:
            st.caption(f"Showing {len(filtered_df)} rules")
            
            # Display columns
            display_cols = ['S.No', 'Business Field', 'Dimension', 'Data Quality Rule', 'Issues Found', 'Issues Found Example']
            display_cols = [c for c in display_cols if c in filtered_df.columns]
            
            # Renumber for display
            filtered_df_display = filtered_df.copy()
            filtered_df_display['S.No'] = range(1, len(filtered_df_display) + 1)
            
            # Style the dataframe
            def highlight_issues(val):
                if isinstance(val, int) and val > 0:
                    return 'background-color: #fee2e2; color: #991b1b; font-weight: bold;'
                return ''
            
            def highlight_examples(val):
                if isinstance(val, str):
                    if '✓ Client provided' in val:
                        return 'background-color: #dcfce7; color: #166534; font-size: 11px; font-weight: bold;'
                    elif val.startswith('✓'):
                        return 'background-color: #dcfce7; color: #166534; font-size: 11px; font-style: italic;'
                    elif len(val) > 0:
                        return 'background-color: #fef3c7; color: #92400e; font-size: 11px;'
                return ''
            
            def highlight_dimension(val):
                colors = {
                    'Accuracy': 'background-color: #dbeafe; color: #1e40af',
                    'Completeness': 'background-color: #dcfce7; color: #166534',
                    'Consistency': 'background-color: #fef3c7; color: #92400e',
                    'Validity': 'background-color: #fce7f3; color: #9d174d',
                    'Uniqueness': 'background-color: #f3e8ff; color: #6b21a8',
                    'Timeliness': 'background-color: #ccfbf1; color: #0f766e',
                    'Integrity': 'background-color: #fee2e2; color: #991b1b',
                    'Conformity': 'background-color: #e0e7ff; color: #3730a3',
                    'Reliability': 'background-color: #ffedd5; color: #9a3412',
                    'Relevance': 'background-color: #ecfccb; color: #3f6212',
                    'Precision': 'background-color: #fae8ff; color: #86198f',
                    'Accessibility': 'background-color: #e0f2fe; color: #075985',
                    'Character Length': 'background-color: #fde68a; color: #92400e; font-weight: bold'
                }
                return colors.get(val, '')
            
            row_count = len(filtered_df_display)
            table_height = min(max(row_count * 35 + 50, 200), 800)
            
            styled_df = filtered_df_display[display_cols].style\
                .applymap(highlight_dimension, subset=['Dimension'])\
                .applymap(highlight_issues, subset=['Issues Found'])\
                .applymap(highlight_examples, subset=['Issues Found Example'])
            
            st.dataframe(styled_df, use_container_width=True, height=table_height)
            
            # Summary charts
            dim_counts = filtered_df['Dimension'].value_counts().reset_index()
            dim_counts.columns = ['Dimension', 'Rule Count']
            
            col1, col2 = st.columns(2)
            with col1:
                fig = px.bar(dim_counts, x='Dimension', y='Rule Count', 
                            color='Dimension', title="Rules by DQ Dimension")
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # Source distribution
                source_counts = filtered_df['Source'].apply(
                    lambda x: 'Client Rules' if 'Client' in str(x) else 'AI Rules'
                ).value_counts().reset_index()
                source_counts.columns = ['Source', 'Count']
                
                fig = px.pie(source_counts, values='Count', names='Source',
                            title="Rules by Source",
                            color_discrete_map={'Client Rules': '#86efac', 'AI Rules': '#93c5fd'})
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("No rules match the current filters")


def _render_match_rules_tab():
    state = st.session_state.app_state
    rules = generate_match_rules(state.df, state.column_profiles)

    if not rules:
        st.warning("No match rules could be generated")
        return

    st.subheader("Suggested Match Rules")
    rules_df = pd.DataFrame(rules)

    def color_prob(val):
        if val in ['Strongest', 'Very Strong', 'Enterprise']: 
            return 'background-color: #10b981; color: white'
        elif val in ['Strong', 'Good']: 
            return 'background-color: #3b82f6; color: white'
        else: 
            return 'background-color: #f59e0b; color: white'

    styled_df = rules_df[['Rule No', 'Rule Type', 'Columns', 'Match Probability', 'Rationale']].style.applymap(
        color_prob, subset=['Match Probability']
    )
    st.dataframe(styled_df, use_container_width=True, hide_index=True)

    st.subheader("Rule Details")
    for rule in rules:
        with st.expander(f"{rule['Rule No']}: {rule['Rule Type']} Match on {rule['Columns']}"):
            st.write(f"**Probability:** {rule['Match Probability']}")
            st.write(f"**Confidence Score:** {rule.get('Confidence', 'N/A')}")
            st.write(f"**Rationale:** {rule['Rationale']}")


def _render_export_tab():
    state = st.session_state.app_state

    st.subheader("📥 Export Profiling Report")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("📊 Generate Excel Report", type="primary", use_container_width=True):
            _generate_excel_report()
    with col2:
        if st.button("📋 Generate JSON Report", type="primary", use_container_width=True):
            _generate_json_report()


# ==========================================
# EXPORT FUNCTIONS
# ==========================================

def _analyze_special_chars_detailed(df):
    data = []
    for col in df.select_dtypes(include=['object']).columns:
        counter = Counter()
        for val in df[col].dropna().astype(str):
            for char in set(val):
                if ord(char) > 127 or (not char.isalnum() and not char.isspace()):
                    counter[char] += val.count(char)
        for char, count in counter.most_common():
            try: 
                uname = unicodedata.name(char)
            except: 
                uname = "UNKNOWN"
            data.append({'Column': col, 'Character': char, 'Unicode Name': uname, 'Count': count})
    return data


def _apply_dq_rules_styling(workbook, worksheet, validation_df):
    """Apply professional styling to Data Quality Rules sheet"""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    dimension_colors = {
        'Accuracy': 'DBEAFE',
        'Completeness': 'DCFCE7',
        'Consistency': 'FEF3C7',
        'Validity': 'FCE7F3',
        'Uniqueness': 'F3E8FF',
        'Timeliness': 'CCFBF1',
        'Integrity': 'FEE2E2',
        'Conformity': 'E0E7FF',
        'Reliability': 'FFEDD5',
        'Relevance': 'ECFCCB',
        'Precision': 'FAE8FF',
        'Accessibility': 'E0F2FE',
        'Character Length': 'FDE68A'
    }
    
    header_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, column_title in enumerate(validation_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_num, row in enumerate(validation_df.itertuples(index=False), 2):
        dimension = row[2] if len(row) > 2 else ''
        issues_found = row[4] if len(row) > 4 else 0
        
        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.border = thin_border
            
            if col_num == 3:
                color = dimension_colors.get(str(dimension), 'FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                cell.font = Font(bold=True)
            
            if col_num == 5:
                try:
                    issue_count = int(issues_found) if isinstance(issues_found, (int, float)) else 0
                    if issue_count > 0:
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                        cell.font = Font(color='991B1B', bold=True)
                    else:
                        cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                        cell.font = Font(color='166534', bold=True)
                except:
                    pass
            
            if col_num == 6:
                if isinstance(value, str):
                    if value.startswith('✓'):
                        cell.fill = PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid')
                        cell.font = Font(color='166534', italic=True, size=10)
                    elif len(value) > 0:
                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                        cell.font = Font(color='92400E', size=10)
            
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    column_widths = {
        'A': 6,
        'B': 20,
        'C': 18,
        'D': 35,
        'E': 12,
        'F': 40
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    worksheet.row_dimensions[1].height = 25
    worksheet.freeze_panes = 'A2'


def _generate_excel_report():
    state = st.session_state.app_state
    
    # Initialize unified rules attributes if they don't exist
    if not hasattr(state, 'unified_rules_generated'):
        state.unified_rules_generated = False
    if not hasattr(state, 'unified_validation_rules'):
        state.unified_validation_rules = None
    
    progress = st.progress(0)
    output = io.BytesIO()

    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df = state.df
            profiles = state.column_profiles
            
            from openpyxl.styles import PatternFill, Font, Alignment
            workbook = writer.book

            progress.progress(10)
            total_missing = sum(p.null_count for p in profiles.values())
            quality = sum(getattr(p, 'non_null_percentage', 100) for p in profiles.values()) / len(profiles) if profiles else 0

            pd.DataFrame({
                'Metric': ['Total Rows', 'Total Columns', 'Total Cells', 'Missing Cells', 'Completeness %', 'Quality Score', 'Generated At'],
                'Value': [len(df), len(df.columns), len(df)*len(df.columns), total_missing, 
                         f"{((len(df)*len(df.columns)-total_missing)/(len(df)*len(df.columns))*100):.2f}%",
                         f"{quality:.1f}%", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            }).to_excel(writer, sheet_name='Executive Summary', index=False)

            progress.progress(25)
            profile_data = []
            for col, p in profiles.items():
                dup_count = p.total_rows - p.unique_count
                profile_data.append({
                    'Column Name': col,
                    'Data Type': p.dtype,
                    'Total Rows': p.total_rows,
                    'Non-Null Count': p.total_rows - p.null_count,
                    'Null Count': p.null_count,
                    'Null Percentage': f"{p.null_percentage:.2f}%",
                    'Unique Count': p.unique_count,
                    'Duplicate Count': dup_count,
                    'Unique Percentage': f"{p.unique_percentage:.2f}%",
                    'Min Length': getattr(p, 'min_length', 'N/A'),
                    'Max Length': getattr(p, 'max_length', 'N/A'),
                    'Avg Length': f"{getattr(p, 'avg_length', 0):.2f}",
                    'Risk Level': getattr(p, 'risk_level', 'Low'),
                    'Risk Score': getattr(p, 'risk_score', 0)
                })
            pd.DataFrame(profile_data).to_excel(writer, sheet_name='Column Profiles', index=False)

            progress.progress(40)
            chars = _analyze_special_chars_detailed(df)
            (pd.DataFrame(chars) if chars else pd.DataFrame({'Message': ['No special characters found']})).to_excel(
                writer, sheet_name='Special Characters', index=False)

            # 4. Match Rule Suggestion
            progress.progress(55)
            match_rules = generate_match_rules(df, profiles)
            pd.DataFrame(match_rules).to_excel(writer, sheet_name='Match Rules', index=False)

            # 5. DQ Dimension Summary
            progress.progress(75)
            progress.progress(90)
            if state.unified_rules_generated and state.unified_validation_rules is not None and not state.unified_validation_rules.empty:
                dq_summary = state.unified_validation_rules.groupby('Dimension').agg({
                    'S.No': 'count',
                    'Issues Found': 'sum'
                }).reset_index()
                dq_summary.columns = ['Dimension', 'Rule Count', 'Total Issues']
                dq_summary = dq_summary.sort_values('Rule Count', ascending=False)
                dq_summary.to_excel(writer, sheet_name='DQ Summary', index=False)

        progress.progress(100)
        
        output.seek(0)
        file_bytes = output.getvalue()
        
        orig_name = getattr(state, 'filename', None) or 'dataset'
        base_name = os.path.splitext(str(orig_name))[0]
        safe_orig = re.sub(r'[^A-Za-z0-9_.-]', '_', base_name)
        export_filename = f"{safe_orig}_Data_Profile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        st.download_button(
            "⬇️ Download Excel Report", 
            data=file_bytes,
            file_name=export_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        show_toast("Excel report generated!", "success")
    except Exception as e:
        st.error(f"Error generating Excel: {str(e)}")


def _generate_json_report():
    import json
    state = st.session_state.app_state
    
    # Initialize unified rules attributes if they don't exist
    if not hasattr(state, 'unified_rules_generated'):
        state.unified_rules_generated = False
    if not hasattr(state, 'unified_validation_rules'):
        state.unified_validation_rules = None

    try:
        df = state.df
        profiles = state.column_profiles

        report = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'ai_powered': True,
                'dq_dimensions': AIValidationEngine.DQ_DIMENSIONS
            },
            'executive_summary': {
                'missing_cells': sum(p.null_count for p in profiles.values()),
                'completeness': f"{sum(getattr(p, 'non_null_percentage', 100) for p in profiles.values())/len(profiles):.2f}%" if profiles else "N/A"
            },
            'column_profiles': {
                col: {
                    'type': p.dtype,
                    'null_percentage': p.null_percentage,
                    'unique_percentage': p.unique_percentage,
                    'duplicate_count': p.total_rows - p.unique_count
                } for col, p in profiles.items()
            },
            'match_rules': generate_match_rules(df, profiles)
        }

        json_data = json.dumps(report, indent=2, default=str)
        
        orig_name = getattr(state, 'filename', None) or 'dataset'
        base_name = os.path.splitext(str(orig_name))[0]
        safe_orig = re.sub(r'[^A-Za-z0-9_.-]', '_', base_name)
        export_filename = f"{safe_orig}_Data_Profile_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

        st.download_button(
            "⬇️ Download JSON Report", 
            data=json_data,
            file_name=export_filename,
            mime="application/json"
        )
        show_toast("JSON report generated!", "success")
    except Exception as e:
        st.error(f"Error generating JSON: {str(e)}")