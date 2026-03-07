"""Comprehensive AI Data Quality Rule Engine for Rule Generator"""

import logging
import re
import json
from typing import Dict, List, Any, Optional
import pandas as pd

logger = logging.getLogger(__name__)


def scan_excel_for_rule_sheets(uploaded_file) -> List[str]:
    """Scan Excel workbook for rule-related sheets
    
    Returns list of sheet names that appear to contain rules
    """
    try:
        import openpyxl
        
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, data_only=False)
        
        # Rule-related sheet name patterns
        rule_patterns = [
            r'rule\s*generator',
            r'validation\s*rules?',
            r'data\s*quality\s*rules?',
            r'dq\s*rules?',
            r'field\s*validation',
            r'business\s*rules?',
            r'validation\s*instruction',
            r'rules?',
            r'validation'
        ]
        
        rule_sheets = []
        for sheet_name in wb.sheetnames:
            for pattern in rule_patterns:
                if re.search(pattern, sheet_name, re.IGNORECASE):
                    rule_sheets.append(sheet_name)
                    logger.info(f"Found rule-related sheet: {sheet_name}")
                    break
        
        wb.close()
        return rule_sheets
        
    except Exception as e:
        logger.error(f"Error scanning for rule sheets: {str(e)}")
        return []


def deep_scan_rule_sheet(uploaded_file_path: str, sheet_name: str, header_row: int = 0) -> Dict[str, Any]:
    """Perform deep scan of a rule sheet to extract all rule information
    
    Args:
        uploaded_file_path: Path to the uploaded Excel file
        sheet_name: Name of the sheet to scan
        header_row: Row index where column headers are located (0-based)
    
    Scans:
    - ALL rows (above and below header row)
    - All columns
    - Cell comments
    - Cell notes
    - Merged cells
    - Hidden rows/columns
    
    Returns dict mapping column names to their detected rules
    """
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(uploaded_file_path, data_only=False)
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return {}
        
        ws = wb[sheet_name]
        
        logger.info(f"=== DEEP SCANNING RULE SHEET: {sheet_name} ===")
        logger.info(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        logger.info(f"Header row specified: {header_row + 1} (1-based)")
        
        # Step 1: Extract column names from header row
        column_names = []
        header_row_idx = header_row + 1  # Convert to 1-based for openpyxl
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            if cell.value:
                column_names.append({
                    'name': str(cell.value).strip(),
                    'col_idx': col_idx,
                    'comment': cell.comment.text if cell.comment else None
                })
                logger.info(f"Column {col_idx}: {cell.value}")
        
        logger.info(f"Found {len(column_names)} columns in header row {header_row_idx}")
        
        # Step 2: Scan ALL rows for rules (above and below header)
        column_rules = {}
        
        for col_info in column_names:
            col_name = col_info['name']
            col_idx = col_info['col_idx']
            rules_found = []
            
            # Add header comment if exists
            if col_info['comment']:
                rules_found.append(f"[Header Comment] {col_info['comment']}")
                logger.info(f"Column '{col_name}': Found header comment")
            
            # Scan rows ABOVE header (rules often written above)
            for row_idx in range(1, header_row_idx):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    value = str(cell.value).strip()
                    if value and len(value) < 200:  # Reasonable rule length
                        rules_found.append(f"[Row {row_idx}] {value}")
                        logger.info(f"Column '{col_name}': Found rule in row {row_idx} above header: {value[:50]}")
                
                # Check for comments in rows above
                if cell.comment:
                    rules_found.append(f"[Row {row_idx} Comment] {cell.comment.text}")
                    logger.info(f"Column '{col_name}': Found comment in row {row_idx}")
            
            # Scan rows BELOW header (first 10 rows for data type hints)
            for row_idx in range(header_row_idx + 1, min(header_row_idx + 11, ws.max_row + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    value = str(cell.value).strip()
                    # Check if this looks like a rule/instruction (not data)
                    if any(keyword in value.lower() for keyword in [
                        'char', 'varchar', 'number', 'numeric', 'decimal', 'date',
                        'mandatory', 'required', 'max', 'length', 'format'
                    ]):
                        rules_found.append(f"[Row {row_idx}] {value}")
                        logger.info(f"Column '{col_name}': Found rule in row {row_idx} below header: {value[:50]}")
                
                # Check for comments in rows below
                if cell.comment:
                    rules_found.append(f"[Row {row_idx} Comment] {cell.comment.text}")
            
            # Scan adjacent cells (left and right) in header row
            # Check cell to the right
            if col_idx < ws.max_column:
                right_cell = ws.cell(row=header_row_idx, column=col_idx + 1)
                if right_cell.value and col_idx + 1 not in [c['col_idx'] for c in column_names]:
                    value = str(right_cell.value).strip()
                    if value and len(value) < 200:
                        rules_found.append(f"[Adjacent Right] {value}")
                        logger.info(f"Column '{col_name}': Found rule in adjacent right cell: {value[:50]}")
            
            # Store all found rules for this column
            if rules_found:
                column_rules[col_name] = ' | '.join(rules_found)
                logger.info(f"Column '{col_name}': Total rules found: {len(rules_found)}")
            else:
                logger.info(f"Column '{col_name}': No rules found, will use AI generation")
        
        wb.close()
        
        logger.info(f"=== SCAN COMPLETE: Found rules for {len(column_rules)}/{len(column_names)} columns ===")
        return column_rules
        
    except Exception as e:
        logger.error(f"Error deep scanning rule sheet: {str(e)}", exc_info=True)
        return {}


def post_process_rules(rules: List[Dict[str, Any]], metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Post-process generated rules to ensure correct dimension mapping and formatting
    
    Applies these transformations:
    1. Reclassify any VARCHAR/CHAR/CHARACTER rules to 'Character Length' dimension
    2. Ensure human-readable formatting for length constraints
    3. Ensure Completeness rules exist for mandatory fields
    4. Ensure separate line items for each dimension
    """
    processed_rules = []
    has_completeness_rule = False
    has_character_length_rule = False
    
    for rule in rules:
        dimension = rule.get('dimension', '')
        rule_text = rule.get('data_quality_rule', '')
        
        # Check if this is a character length rule based on keywords
        char_keywords = ['CHAR', 'CHARACTER', 'VARCHAR', 'VARCHAR2', 'STRING', 'TEXT', 
                        'maximum', 'max length', 'characters', 'character length']
        
        is_length_rule = any(keyword.lower() in rule_text.lower() for keyword in char_keywords)
        
        # Reclassify to Character Length if it contains length-related keywords
        if is_length_rule and 'character' in rule_text.lower():
            rule['dimension'] = 'Character Length'
            has_character_length_rule = True
            
            # Ensure proper formatting: "should be maximum X characters"
            if metadata.get('max_length'):
                field_name = rule.get('business_field', '')
                rule['data_quality_rule'] = f"{field_name} should be maximum {metadata['max_length']} characters"
        
        # Check if this is a completeness rule
        if dimension == 'Completeness' or 'must not be blank' in rule_text.lower():
            has_completeness_rule = True
            
            # Ensure proper formatting: "must not be blank"
            if 'must not be blank' not in rule_text.lower():
                field_name = rule.get('business_field', '')
                rule['data_quality_rule'] = f"{field_name} must not be blank"
        
        processed_rules.append(rule)
    
    # Ensure Completeness rule exists for mandatory fields
    if metadata.get('mandatory') and not has_completeness_rule:
        field_name = processed_rules[0].get('business_field', '') if processed_rules else ''
        completeness_rule = {
            'business_field': field_name,
            'dimension': 'Completeness',
            'data_quality_rule': f"{field_name} must not be blank",
            'issues_found': 0,
            'issues_found_example': '✓ All values valid - No issues found'
        }
        processed_rules.append(completeness_rule)
    
    # Ensure Character Length rule exists if max_length is detected
    if metadata.get('max_length') and not has_character_length_rule:
        field_name = processed_rules[0].get('business_field', '') if processed_rules else ''
        char_length_rule = {
            'business_field': field_name,
            'dimension': 'Character Length',
            'data_quality_rule': f"{field_name} should be maximum {metadata['max_length']} characters",
            'issues_found': 0,
            'issues_found_example': '✓ All values valid - No issues found'
        }
        processed_rules.append(char_length_rule)
    
    return processed_rules


def extract_column_rules_from_sheet_data(rule_data: Dict, worksheet) -> Dict[str, str]:
    """Extract column-to-rule mappings from scanned sheet data
    
    Attempts to identify:
    - Column names (usually in first few rows)
    - Associated rules (in adjacent cells, comments, or same row)
    """
    column_rules = {}
    
    # Strategy 1: Look for header row (first 5 rows)
    potential_headers = []
    for cell_ref, data in rule_data.items():
        if data['row'] <= 5 and data['value']:
            # Check if this looks like a column name
            value = data['value']
            if len(value) < 100 and not value.startswith('='):  # Not a formula
                potential_headers.append({
                    'name': value,
                    'row': data['row'],
                    'col': data['col'],
                    'cell_ref': cell_ref
                })
    
    # Strategy 2: For each potential header, look for rules in adjacent cells or comments
    for header in potential_headers:
        column_name = header['name']
        rules_found = []
        
        # Check same row (to the right)
        for cell_ref, data in rule_data.items():
            if data['row'] == header['row'] and data['col'] > header['col']:
                if data['value'] and len(data['value']) > 10:  # Likely a rule description
                    rules_found.append(data['value'])
        
        # Check rows below (same column)
        for cell_ref, data in rule_data.items():
            if data['col'] == header['col'] and data['row'] > header['row'] and data['row'] <= header['row'] + 3:
                if data['value']:
                    rules_found.append(data['value'])
        
        # Check for comments on header cell
        if rule_data[header['cell_ref']].get('comment'):
            rules_found.append(rule_data[header['cell_ref']]['comment'])
        
        if rules_found:
            column_rules[column_name] = ' | '.join(rules_found)
            logger.info(f"Mapped column '{column_name}' to rules: {column_rules[column_name][:100]}")
    
    return column_rules


def extract_comprehensive_metadata(column_name: str, rule_text: str = None) -> Dict[str, Any]:
    """Extract comprehensive metadata from column names and rule text
    
    Scans both column name and associated rule text for patterns
    """
    metadata = {
        'max_length': None,
        'data_type_hint': None,
        'mandatory': False,
        'format_restrictions': [],
        'precision': None,
        'scale': None,
        'uniqueness_required': False,
        'conditional_rule': None,
        'allowed_values': None,
        'detected_raw_text': rule_text or '',
        'extracted_pattern': [],
        'conflict_flag': False,
        'original_name': column_name
    }
    
    # Combine column name and rule text for scanning
    scan_text = f"{column_name} {rule_text or ''}"
    
    # Check for mandatory indicator (asterisk)
    if '*' in scan_text:
        metadata['mandatory'] = True
        metadata['extracted_pattern'].append('mandatory (*)')
        logger.info(f"Column '{column_name}' marked as mandatory (contains *)")
    
    # Check for mandatory keywords
    if re.search(r'\b(mandatory|required|compulsory|must|not\s*null|cannot\s*be\s*null|must\s*be\s*provided)\b', scan_text, re.IGNORECASE):
        metadata['mandatory'] = True
        metadata['extracted_pattern'].append('mandatory (keyword)')
        logger.info(f"Column '{column_name}' marked as mandatory (keyword found)")
    
    # Check for uniqueness
    if re.search(r'\b(unique|distinct|pk|primary\s*key|no\s*duplicate)\b', scan_text, re.IGNORECASE):
        metadata['uniqueness_required'] = True
        metadata['extracted_pattern'].append('unique')
        logger.info(f"Column '{column_name}' marked as unique")
    
    # Pattern 1: VARCHAR2(360), VARCHAR(26), CHAR(19), CHARACTERS(21), STRING(50), TEXT(100)
    pattern1 = r'(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?'
    matches = list(re.finditer(pattern1, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            dtype = match.group(1).upper()
            length = int(match.group(2))
            if metadata['max_length'] and metadata['max_length'] != length:
                metadata['conflict_flag'] = True
                logger.warning(f"Conflict detected for '{column_name}': multiple lengths found")
            metadata['data_type_hint'] = dtype
            metadata['max_length'] = length
            metadata['extracted_pattern'].append(f"{dtype}({length})")
            logger.info(f"Extracted from '{column_name}': Type={dtype}, MaxLength={length}")
    
    # Pattern 2: NUMBER(10,2), DECIMAL(15,2), NUMERIC(8,2)
    pattern2 = r'(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?'
    matches = list(re.finditer(pattern2, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            dtype = match.group(1).upper()
            precision = int(match.group(2))
            scale = int(match.group(3)) if match.group(3) else None
            metadata['data_type_hint'] = dtype
            metadata['precision'] = precision
            metadata['scale'] = scale
            pattern_str = f"{dtype}({precision},{scale})" if scale else f"{dtype}({precision})"
            metadata['extracted_pattern'].append(pattern_str)
            logger.info(f"Extracted from '{column_name}': Type={dtype}, Precision={precision}, Scale={scale}")
    
    # Pattern 3: NUM(10), NUM 10, NUMBER 10, INT(5)
    pattern3 = r'(NUM|NUMBER|INT|INTEGER)\s*\(?\s*(\d+)\s*\)?'
    if not metadata['precision']:  # Only if not already found
        matches = list(re.finditer(pattern3, scan_text, re.IGNORECASE))
        if matches:
            for match in matches:
                metadata['data_type_hint'] = 'NUMBER'
                metadata['precision'] = int(match.group(2))
                metadata['extracted_pattern'].append(f"NUMBER({metadata['precision']})")
                logger.info(f"Extracted from '{column_name}': Type=NUMBER, Precision={metadata['precision']}")
    
    # Pattern 4: Just numbers in parentheses like "Name (360)" or "Code (19)"
    pattern4 = r'\((\d+)\)'
    if not metadata['max_length']:  # Only if not already found
        matches = list(re.finditer(pattern4, scan_text))
        if matches:
            metadata['max_length'] = int(matches[0].group(1))
            metadata['extracted_pattern'].append(f"length({metadata['max_length']})")
            logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']}")
    
    # Pattern 5: "max length 50", "maximum 30 characters", "max 100"
    pattern5 = r'(?:max(?:imum)?\s*(?:length|len|size|chars?)?)\s*[:\-]?\s*(\d+)'
    matches = list(re.finditer(pattern5, scan_text, re.IGNORECASE))
    if matches:
        for match in matches:
            length = int(match.group(1))
            if metadata['max_length'] and metadata['max_length'] != length:
                metadata['conflict_flag'] = True
            metadata['max_length'] = length
            metadata['extracted_pattern'].append(f"max length {length}")
            logger.info(f"Extracted from '{column_name}': MaxLength={length} (from max length pattern)")
    
    # Pattern 6: Format restrictions
    if re.search(r'\b(uppercase|upper|caps|all\s*caps)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('UPPERCASE')
        metadata['extracted_pattern'].append('UPPERCASE')
    if re.search(r'\b(lowercase|lower)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('LOWERCASE')
        metadata['extracted_pattern'].append('LOWERCASE')
    if re.search(r'\b(alphanumeric|alpha\s*numeric|no\s*special\s*char)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('ALPHANUMERIC')
        metadata['extracted_pattern'].append('ALPHANUMERIC')
    if re.search(r'\b(no\s*spaces?|trim)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_SPACES')
        metadata['extracted_pattern'].append('NO_SPACES')
    if re.search(r'\b(no\s*leading\s*zero)\b', scan_text, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_LEADING_ZERO')
        metadata['extracted_pattern'].append('NO_LEADING_ZERO')
    
    # Pattern 7: Value restrictions (Yes/No, Y/N, etc.)
    if re.search(r'\b(yes\s*/\s*no|y\s*/\s*n|true\s*/\s*false)\b', scan_text, re.IGNORECASE):
        metadata['allowed_values'] = ['Yes', 'No']
        metadata['extracted_pattern'].append('Yes/No')
        logger.info(f"Extracted allowed values from '{column_name}': Yes/No")
    
    # Pattern 8: List of Values, LOV
    if re.search(r'\b(list\s*of\s*values?|lov|enum)\b', scan_text, re.IGNORECASE):
        metadata['extracted_pattern'].append('LOV')
        logger.info(f"Column '{column_name}' has List of Values restriction")
    
    # Pattern 9: Conditional rules
    conditional_match = re.search(r'(mandatory\s*if|required\s*when|only\s*if|depends\s*on)\s+(.+?)(?:\.|$)', scan_text, re.IGNORECASE)
    if conditional_match:
        metadata['conditional_rule'] = conditional_match.group(0)
        metadata['extracted_pattern'].append(f"conditional: {conditional_match.group(1)}")
        logger.info(f"Extracted conditional rule from '{column_name}': {metadata['conditional_rule']}")
    
    # Pattern 10: Date formats
    date_patterns = [
        (r'DD-MM-YYYY', 'DD-MM-YYYY'),
        (r'YYYY-MM-DD', 'YYYY-MM-DD'),
        (r'MM/DD/YYYY', 'MM/DD/YYYY'),
        (r'past\s*date\s*only', 'past date only'),
        (r'future\s*date\s*not\s*allowed', 'no future date'),
        (r'system\s*date', 'system date')
    ]
    for pattern, label in date_patterns:
        if re.search(pattern, scan_text, re.IGNORECASE):
            metadata['extracted_pattern'].append(label)
            logger.info(f"Extracted date pattern from '{column_name}': {label}")
    
    if metadata['format_restrictions']:
        logger.info(f"Extracted format restrictions from '{column_name}': {metadata['format_restrictions']}")
    
    return metadata


def generate_comprehensive_ai_prompt(column_name: str, sample_data: List[str], data_type: str,
                                     null_pct: float, unique_pct: float, 
                                     metadata: Dict[str, Any], rule_source: str) -> str:
    """Generate comprehensive AI prompt for rule generation with exact output format"""
    
    samples = sample_data[:100] if sample_data else []
    sample_str = json.dumps(samples, indent=2)
    
    # Build metadata section
    metadata_section = ""
    if metadata.get('detected_raw_text'):
        metadata_section += f"\n=== RULES FROM EXISTING SHEET ===\n{metadata['detected_raw_text']}\n"
    
    if metadata.get('extracted_pattern'):
        metadata_section += f"\n=== EXTRACTED PATTERNS ===\n"
        for pattern in metadata['extracted_pattern']:
            metadata_section += f"- {pattern}\n"
    
    if metadata.get('max_length'):
        metadata_section += f"\n- Maximum Length: {metadata['max_length']} characters\n"
    if metadata.get('data_type_hint'):
        metadata_section += f"- Data Type Hint: {metadata['data_type_hint']}\n"
    if metadata.get('precision'):
        metadata_section += f"- Numeric Precision: {metadata['precision']}\n"
    if metadata.get('scale'):
        metadata_section += f"- Numeric Scale: {metadata['scale']}\n"
    if metadata.get('mandatory'):
        metadata_section += f"- Mandatory Field: YES\n"
    if metadata.get('uniqueness_required'):
        metadata_section += f"- Uniqueness Required: YES\n"
    if metadata.get('format_restrictions'):
        metadata_section += f"- Format Restrictions: {', '.join(metadata['format_restrictions'])}\n"
    if metadata.get('allowed_values'):
        metadata_section += f"- Allowed Values: {', '.join(metadata['allowed_values'])}\n"
    if metadata.get('conditional_rule'):
        metadata_section += f"- Conditional Rule: {metadata['conditional_rule']}\n"
    if metadata.get('conflict_flag'):
        metadata_section += f"- ⚠️ CONFLICT DETECTED: Multiple conflicting patterns found\n"
    
    prompt = f"""You are an AI Data Quality Rule Engine.

=== COLUMN INFORMATION ===
Column Name: {column_name}
Rule Source: {rule_source}
Detected Data Type: {data_type}
Sample Values: {sample_str}
Null Percentage: {null_pct:.1f}%
Unique Percentage: {unique_pct:.1f}%
{metadata_section}

=== YOUR TASK ===

Generate validation rules in the EXACT format shown below.

Based on the Rule Source:

If Rule Source = "Rules from Existing Sheet":
- Use the extracted patterns and rules from the sheet
- Interpret them logically
- Generate validation rules based on detected patterns

If Rule Source = "Generated by AI":
- No rules found in existing sheets
- Generate intelligent validation rules based on:
  * Column name semantics
  * Sample data patterns
  * Business meaning
  * Enterprise data standards

=== PATTERN INTERPRETATION ===

Length Patterns:
- "30 Characters" → Data Type = String, Max Length = 30
- "varchar2(150)" → Data Type = String, Max Length = 150
- "char(20)" → Data Type = String, Max Length = 20

Numeric Patterns:
- "Number" → Data Type = Numeric
- "number(10)" → Numeric, Max digits = 10
- "numeric(8,2)" → Numeric, Total digits = 8, Decimal = 2

Date Patterns:
- "Date" → Data Type = Date
- "DD-MM-YYYY" → Date format specified

=== DIMENSION CLASSIFICATION (CRITICAL RULES) ===

**MANDATORY DIMENSION MAPPING:**

1. **Character Length** - Use for ANY string length constraint containing these keywords:
   - CHAR, CHARACTER, VARCHAR, VARCHAR2, STRING, TEXT
   - "max length", "maximum characters", "length constraint"
   - Example: VARCHAR2(30) → Dimension = "Character Length"
   - Example: CHAR(19) → Dimension = "Character Length"
   - ⚠️ IMPORTANT: Even if traditionally "Conformity", these MUST be "Character Length"

2. **Completeness** - Use for mandatory/required fields:
   - Fields marked with * (asterisk)
   - Keywords: mandatory, required, not null, must not be blank
   - Rule format: "[Field Name] must not be blank"

3. **Validity** - Use for data type and format validation:
   - "must be numeric", "must be date", "must be valid email"
   - Data type checks, format pattern validation

4. **Uniqueness** - Use for unique constraints:
   - unique, distinct, no duplicates, primary key

5. **Conformity** - Use for format restrictions (NOT length):
   - UPPERCASE, lowercase, alphanumeric only
   - Specific patterns, regex validation

6. **Accuracy** - Use for value ranges and precision:
   - Value ranges, decimal precision, scale

=== HUMAN-READABLE RULE FORMATTING ===

Transform technical constraints to natural language:

Technical Input → Human-Readable Output:
- "VARCHAR2(30)" → "Account Name should be maximum 30 characters"
- "CHAR(19)" → "Code should be maximum 19 characters"
- "VARCHAR(100)" → "Description should be maximum 100 characters"
- "Mandatory *" → "Field Name must not be blank"
- "NUMBER(10,2)" → "Amount must be numeric with maximum 10 digits and 2 decimal places"

=== RULE STATEMENT FORMAT ===

Write rules in this format: [Field Name] + must/should + condition

Examples:
- "Account Name should be maximum 30 characters" (Character Length)
- "Account Name must not be blank" (Completeness)
- "Interface Line Number must be numeric" (Validity)
- "Date Placed in Service must be in YYYY/MM/DD date format" (Validity)
- "Asset Type should be maximum 11 characters" (Character Length)

=== IMPORTANT GUIDELINES ===

1. **For Character Length dimension**: ALWAYS use "should be maximum X characters"
2. **For Completeness dimension**: ALWAYS use "must not be blank" for mandatory fields
3. **For Validity dimension**: Use "must be numeric" or "must be in [format]"
4. **For Uniqueness**: Use "must be unique"

5. Generate SEPARATE rules for each dimension:
   - If field has VARCHAR2(30) AND is mandatory → Generate TWO rules:
     * Rule 1: Dimension = "Character Length", Rule = "[Field] should be maximum 30 characters"
     * Rule 2: Dimension = "Completeness", Rule = "[Field] must not be blank"

6. PRIORITY: For every mandatory field, ALWAYS generate a Completeness rule
7. Use detected patterns if available
8. Be dynamic - no hardcoded values
9. Set issues_found to 0 if validation passes
10. Set issues_found_example to "✓ All values valid - No issues found" if no issues

=== OUTPUT FORMAT (STRICT JSON) ===

Return an array of rule objects (one per dimension), wrapped in a "rules" key:

{{
  "rules": [
    {{
      "business_field": "{column_name}",
      "dimension": "Character Length",
      "data_quality_rule": "{column_name} should be maximum X characters",
      "issues_found": 0,
      "issues_found_example": "✓ All values valid - No issues found"
    }},
    {{
      "business_field": "{column_name}",
      "dimension": "Completeness",
      "data_quality_rule": "{column_name} must not be blank",
      "issues_found": 0,
      "issues_found_example": "✓ All values valid - No issues found"
    }}
  ]
}}

Return ONLY valid JSON, no markdown."""
    
    return prompt

    """Extract comprehensive metadata from column names with advanced pattern detection
    
    Detects:
    - Length patterns: VARCHAR2(360), CHAR(19), max length 50
    - Numeric patterns: NUMBER(10,2), DECIMAL(15,2), NUM(10)
    - Mandatory indicators: *, mandatory, required
    - Format restrictions: UPPERCASE, ALPHANUMERIC, NO_SPACES
    - Uniqueness: unique, distinct, PK
    - Conditional rules: "if", "when", "depends on"
    
    Returns comprehensive metadata dict
    """
    metadata = {
        'max_length': None,
        'data_type_hint': None,
        'mandatory': False,
        'format_restrictions': [],
        'precision': None,
        'scale': None,
        'uniqueness_required': False,
        'conditional_rule': None,
        'allowed_values': None,
        'original_name': column_name
    }
    
    # Check for mandatory indicator (asterisk)
    if '*' in column_name:
        metadata['mandatory'] = True
        logger.info(f"Column '{column_name}' marked as mandatory (contains *)")
    
    # Check for mandatory keywords
    if re.search(r'\b(mandatory|required|compulsory|must|not\s*null)\b', column_name, re.IGNORECASE):
        metadata['mandatory'] = True
        logger.info(f"Column '{column_name}' marked as mandatory (keyword found)")
    
    # Check for uniqueness
    if re.search(r'\b(unique|distinct|pk|primary\s*key|no\s*duplicate)\b', column_name, re.IGNORECASE):
        metadata['uniqueness_required'] = True
        logger.info(f"Column '{column_name}' marked as unique")
    
    # Pattern 1: VARCHAR2(360), VARCHAR(26), CHAR(19), CHARACTERS(21), STRING(50), TEXT(100)
    pattern1 = r'(VARCHAR2?|CHAR(?:ACTERS)?|STRING|TEXT)\s*\(?\s*(\d+)\s*(?:CHAR|BYTE)?\s*\)?'
    match = re.search(pattern1, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['max_length'] = int(match.group(2))
        logger.info(f"Extracted from '{column_name}': Type={metadata['data_type_hint']}, MaxLength={metadata['max_length']}")
        return metadata
    
    # Pattern 2: NUMBER(10,2), DECIMAL(15,2), NUMERIC(8,2)
    pattern2 = r'(NUMBER|DECIMAL|NUMERIC|FLOAT)\s*\(?\s*(\d+)(?:,\s*(\d+))?\s*\)?'
    match = re.search(pattern2, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = match.group(1).upper()
        metadata['precision'] = int(match.group(2))
        if match.group(3):
            metadata['scale'] = int(match.group(3))
        logger.info(f"Extracted from '{column_name}': Type={metadata['data_type_hint']}, Precision={metadata['precision']}, Scale={metadata.get('scale')}")
        return metadata
    
    # Pattern 3: NUM(10), NUM 10, NUMBER 10, INT(5)
    pattern3 = r'(NUM|NUMBER|INT|INTEGER)\s*\(?\s*(\d+)\s*\)?'
    match = re.search(pattern3, column_name, re.IGNORECASE)
    if match:
        metadata['data_type_hint'] = 'NUMBER'
        metadata['precision'] = int(match.group(2))
        logger.info(f"Extracted from '{column_name}': Type=NUMBER, Precision={metadata['precision']}")
        return metadata
    
    # Pattern 4: Just numbers in parentheses like "Name (360)" or "Code (19)"
    pattern4 = r'\((\d+)\)'
    match = re.search(pattern4, column_name)
    if match:
        metadata['max_length'] = int(match.group(1))
        logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']}")
    
    # Pattern 5: "max length 50", "maximum 30 characters", "max 100"
    pattern5 = r'(?:max(?:imum)?\s*(?:length|len|size|chars?)?)\s*[:\-]?\s*(\d+)'
    match = re.search(pattern5, column_name, re.IGNORECASE)
    if match:
        metadata['max_length'] = int(match.group(1))
        logger.info(f"Extracted from '{column_name}': MaxLength={metadata['max_length']} (from max length pattern)")
    
    # Pattern 6: Format restrictions
    if re.search(r'\b(uppercase|upper|caps|all\s*caps)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('UPPERCASE')
    if re.search(r'\b(lowercase|lower)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('LOWERCASE')
    if re.search(r'\b(alphanumeric|alpha\s*numeric|no\s*special\s*char)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('ALPHANUMERIC')
    if re.search(r'\b(no\s*spaces?|trim)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_SPACES')
    if re.search(r'\b(no\s*leading\s*zero)\b', column_name, re.IGNORECASE):
        metadata['format_restrictions'].append('NO_LEADING_ZERO')
    
    # Pattern 7: Value restrictions (Yes/No, Y/N, etc.)
    if re.search(r'\b(yes\s*/\s*no|y\s*/\s*n|true\s*/\s*false)\b', column_name, re.IGNORECASE):
        metadata['allowed_values'] = ['Yes', 'No']
        logger.info(f"Extracted allowed values from '{column_name}': Yes/No")
    
    # Pattern 8: Conditional rules
    if re.search(r'\b(if|when|depends\s*on)\b', column_name, re.IGNORECASE):
        conditional_match = re.search(r'(if|when|depends\s*on)\s+(.+)', column_name, re.IGNORECASE)
        if conditional_match:
            metadata['conditional_rule'] = conditional_match.group(0)
            logger.info(f"Extracted conditional rule from '{column_name}': {metadata['conditional_rule']}")
    
    if metadata['format_restrictions']:
        logger.info(f"Extracted format restrictions from '{column_name}': {metadata['format_restrictions']}")
    
    return metadata


def generate_comprehensive_ai_prompt(column_name: str, sample_data: List[str], data_type: str,
                                     null_pct: float, unique_pct: float, 
                                     metadata: Dict[str, Any], rule_source: str = None) -> str:
    """Generate comprehensive AI prompt for deep workbook analysis"""
    
    samples = sample_data[:100] if sample_data else []
    sample_str = json.dumps(samples, indent=2)
    
    # Build metadata section
    metadata_section = ""
    if metadata:
        metadata_section = "\n=== EXTRACTED METADATA FROM COLUMN NAME ===\n"
        if metadata.get('max_length'):
            metadata_section += f"- Maximum Length: {metadata['max_length']} characters\n"
        if metadata.get('data_type_hint'):
            metadata_section += f"- Data Type Hint: {metadata['data_type_hint']}\n"
        if metadata.get('precision'):
            metadata_section += f"- Numeric Precision: {metadata['precision']}\n"
        if metadata.get('scale'):
            metadata_section += f"- Numeric Scale: {metadata['scale']}\n"
        if metadata.get('mandatory'):
            metadata_section += f"- Mandatory Field: YES\n"
        if metadata.get('uniqueness_required'):
            metadata_section += f"- Uniqueness Required: YES\n"
        if metadata.get('format_restrictions'):
            metadata_section += f"- Format Restrictions: {', '.join(metadata['format_restrictions'])}\n"
        if metadata.get('allowed_values'):
            metadata_section += f"- Allowed Values: {', '.join(metadata['allowed_values'])}\n"
        if metadata.get('conditional_rule'):
            metadata_section += f"- Conditional Rule: {metadata['conditional_rule']}\n"
    
    if rule_source:
        metadata_section += f"\n=== EXCEL NOTES/COMMENTS/INSTRUCTIONS ===\n{rule_source}\n"
    
    prompt = f"""You are an AI Data Quality Rule Engine.
Your task is to deeply analyze the provided Excel column data dynamically.

=== COLUMN INFORMATION ===
Column Name: {column_name}
Detected Data Type: {data_type}
Sample Values: {sample_str}
Null Percentage: {null_pct:.1f}%
Unique Percentage: {unique_pct:.1f}%
{metadata_section}

=== YOUR RESPONSIBILITIES ===

STEP 1: DETECT CLIENT-DEFINED RULES
Scan the column name, metadata, and Excel notes for patterns such as:

Length Patterns:
- char(20), char 20, characters 20
- varchar2(230), VARCHAR2 221, VARCHAR(100)
- text(100), string(50)
- max length 50, maximum 30 characters

Numeric Patterns:
- number(10), num 10, num(12)
- numeric(8,2), decimal(10,2)
- integer, whole number, int

Date Patterns:
- DD-MM-YYYY, YYYY-MM-DD, MM/DD/YYYY
- Date format, Only past date, Not future date
- Valid date, Date range

Mandatory Indicators:
- * (asterisk in column name)
- mandatory, required, must be provided
- cannot be null, compulsory, not null

Value Restrictions:
- Yes/No, Y/N, True/False
- List of Values, LOV, Enum
- Specific values mentioned
- Reference table mentioned

Format Restrictions:
- Uppercase only, UPPER, ALL CAPS
- Lowercase only, lower
- No special characters, alphanumeric only
- Trim spaces, no leading/trailing spaces
- No leading zero
- Specific regex pattern

Uniqueness:
- unique, distinct, no duplicates
- primary key, PK, identifier

Conditional Rules:
- "Mandatory if [condition]"
- "Required when [condition]"
- Depends on another column

STEP 2: INTERPRET RULES LOGICALLY
If pattern detected:
- char(20) → Max length = 20, Data type = String
- varchar2(230) → Max length = 230, String
- number(10) → Numeric, Max digits = 10
- numeric(8,2) → Total digits 8, 2 decimals
- * in name → Mandatory field
- "Yes/No" → Allowed values: Yes, No

Do NOT hardcode any values. Always extract dynamically from detected pattern.

STEP 3: IF NO CLIENT RULE FOUND
Intelligently infer rules based on:
- Column name semantics
- Sample values (if available)
- Business meaning
- Common enterprise standards

Examples:
- Email → Valid email format, contains @
- Phone/Mobile → Numeric, length 10-15
- Amount/Price/Cost → Numeric, decimal allowed, non-negative
- Date → Valid date format, reasonable range
- Name → Alphabetic with spaces, reasonable length
- ID/Code → Alphanumeric, specific format
- Address → Text, reasonable length
- Percentage → Numeric, 0-100 range
- URL → Valid URL format
- ZIP/Postal → Alphanumeric, specific length

STEP 4: GENERATE STRUCTURED OUTPUT
Return a JSON object with this EXACT structure:

{{
  "business_field_name": "Human-friendly field name",
  "detected_client_rule": "Any rule found in metadata/notes/column name (or 'None')",
  "extracted_pattern": "Pattern detected (e.g., 'VARCHAR2(360)', 'number(10,2)', 'mandatory *')",
  "interpreted_data_type": "String/Numeric/Date/Boolean/etc.",
  "max_length": "Number or null",
  "precision_scale": "For numeric: '10,2' or null",
  "mandatory": "Yes/No based on * or mandatory keyword",
  "allowed_values": "List of allowed values if restricted (or null)",
  "format_restrictions": "Uppercase/Lowercase/Alphanumeric/etc. (or null)",
  "uniqueness_required": "Yes/No",
  "conditional_rule": "Any conditional logic (or null)",
  "rules": [
    {{
      "dimension": "One of: Accuracy, Completeness, Consistency, Validity, Uniqueness, Timeliness, Integrity, Conformity, Reliability, Relevance, Precision, Accessibility",
      "rule_statement": "Human readable rule: [Field Name] + Must/Should + Business Condition"
    }}
  ]
}}

=== RULE WRITING GUIDELINES ===
Write rules in this format: [Field Name] + Must/Should + Business Condition

Priority Order:
1. Client-defined rules from metadata/notes (HIGHEST PRIORITY)
2. Pattern-based rules from column name
3. AI-inferred rules from column semantics
4. Industry standard rules

Examples:
- "Supplier Name must not exceed 360 characters" (from VARCHAR2(360))
- "Invoice Date must be a valid date in DD-MM-YYYY format" (from date pattern)
- "Email Address must follow standard email format" (AI inferred)
- "Tax ID must be alphanumeric with no special characters" (from pattern)
- "Amount must be numeric with maximum 2 decimal places" (from numeric(10,2))
- "Status must be one of: Active, Inactive, Pending" (from LOV)
- "Supplier Name is mandatory and cannot be null" (from * indicator)

=== IMPORTANT CONSTRAINTS ===
- No hardcoded limits without evidence
- If conflicting rules found → mention in detected_client_rule
- If rule found in comment/note → treat as HIGHEST priority
- Be dynamic - extract, don't assume
- Generate 2-4 rules per column minimum
- Always include Completeness rule if mandatory
- Always include Conformity rule if length/format specified
- Always include Validity rule for data type validation

Return ONLY valid JSON, no markdown, no explanation."""
    
    return prompt
